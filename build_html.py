# -*- coding: utf-8 -*-
"""Экспорт Excel + AmoCRM → initial-data.js"""
import json, os, sys, importlib.util

BASE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(BASE)

spec = importlib.util.spec_from_file_location("build_xlsx", os.path.join(ROOT, "build_itmen_v2.py"))
build_xlsx = importlib.util.module_from_spec(spec)
spec.loader.exec_module(build_xlsx)

import openpyxl

XLSX = build_xlsx.OUTPUT
wb = openpyxl.load_workbook(XLSX, data_only=True)

with open(os.path.join(ROOT, "amocrm_analysis.json"), encoding="utf-8") as f:
    AMO = json.load(f)

STAGE_ORDER = [
    "Взят в работу", "Встреча состоялась", "Интерес  Выявлен",
    "Подготовка Пилота", "Пилот", "Пилот Окончен",
    "Предложение выслано", "Согласование бюджета", "Финальный компред",
    "Условия согласованы", "Документы подписаны", "Отгружен",
    "Успешно реализовано", "На паузе",
]

LISTS = {
    "stages": [s for s in STAGE_ORDER if s in AMO["Этап сделки"]],
    "dealTypes": ["Текущий пайплайн"],
    "budgetStatus": ["Подтверждён", "В процессе согласования", "Планируется согласование", "Нет бюджета", "Неизвестно"],
    "categories": ["Горячая", "Тёплая", "Наблюдение", "Отказ"],
    "owners": [
        "Аркадий Мерлейн", "Арслан Ахметшин",
        "Александр Сироткин", "Алексей Кулагин",
    ],
    "industries": sorted(set(AMO["Отрасль"])) + ["Не определена"],
    "budgetPeriods": [
        "Q3 2026", "Q4 2026",
        "Q1 2027", "Q2 2027", "Q3 2027", "Q4 2027",
        "Не определён", "После 2027",
    ],
    "partners": sorted(set(AMO.get("Партнер", []))) + ["Нет партнёра", "Другой"],
    "dml": ["CIO", "CISO", "IT-директор", "Руководитель ITSM", "Не определён"],
}

def deal_from_row(r):
    r = list(r) + [None] * 45
    return {
        "id": r[0], "customer": r[1], "industry": r[2] or "Не определена",
        "owner": r[3] or "Аркадий Мерлейн", "stage": r[4],
        "dealType": r[5] or "Текущий пайплайн", "amount": r[6] or 0,
        "expectedBudget": r[12] or 0,
        "partner": "Нет партнёра", "partnerDiscount": 0, "clientDiscount": 0,
        "manualProb": r[7] or 0,
        "taskDue": str(r[9])[:10] if r[9] else "",
        "budgetPeriod": r[10] or "Не определён",
        "budgetStatus": r[11] or "Неизвестно", "budgetAmount": r[12] or 0,
        "stack": r[13] or "", "competitors": r[14] or "", "complexProject": r[15] or "",
        "pains": r[16] or "", "capabilities": r[17] or "", "dml": r[18] or "Не определён",
        "scores": {
            "loyalty": r[19] or 0, "budget": r[20] or 0, "fit": r[21] or 0,
            "timing": r[22] or 0, "competitive": r[23] or 0, "access": r[24] or 0,
            "technical": r[25] or 0, "commercial": r[26] or 0,
        },
        "scoreReasons": {}, "scoreHistory": [], "scoresOverridden": {},
        "nextStepType": "discovery", "nextStepComment": r[30] or "",
        "riskType": "no_budget" if "бюджет" in str(r[29] or "").lower() else "none",
        "riskComment": r[29] or "",
        "commitStatus": "verbal" if r[32] == "Устное" else "none",
        "lastUpdate": str(r[33])[:10] if r[33] else "",
        "amoId": None,
        "techResearch": default_tech_research(),
    }

def default_tech_research():
    return {
        "classEntries": {},
        "pilotMode": "not_started",
        "pilotPartners": "",
        "integrationNotes": "",
        "searchComment": "",
        "projectCompliancePct": None,
        "pilotCompliancePct": None,
    }

DEALS = []
for r in wb["Паспорт_сделок"].iter_rows(min_row=2, values_only=True):
    if not r[0]:
        continue
    d = deal_from_row(r)
    # migrate legacy commit labels
    legacy = str(r[32] or "")
    mapping = {"Нет": "none", "Устное": "verbal", "Email": "email", "Протокол": "protocol",
               "LOI": "loi", "Гарантийное письмо": "guarantee", "Контракт": "contract"}
    d["commitStatus"] = mapping.get(legacy, d["commitStatus"])
    if r[9]:
        d["taskDue"] = str(r[9])[:10]
    if r[10]:
        d["budgetPeriod"] = str(r[10])
    if r[13]:
        d["techResearch"]["currentSolutions"] = str(r[13])
    if r[14]:
        d["techResearch"]["reviewedProducts"] = [
            {"name": s.strip(), "status": "evaluating", "rejectedReasons": [],
             "appealReasons": [], "openSource": "unknown", "origin": "unknown", "comment": ""}
            for s in str(r[14]).split(",") if s.strip()
        ]
    DEALS.append(d)

SCORING = []
for r in wb["Модель_скоринга"].iter_rows(min_row=4, max_row=11, values_only=True):
    if r[0] and r[0] != "Пороги категорий":
        SCORING.append({
            "name": r[0], "col": r[1], "weight": r[2],
            "s5": r[3], "s3": r[4], "s1": r[5], "s0": r[6], "owner": r[7],
        })

INITIAL = {
    "lists": LISTS,
    "deals": DEALS,
    "scoring": SCORING,
    "pipelineFocus": {
        "title": "Текущий пайплайн — активен",
        "goal": "Коммиты клиентов, даты бюджета, top-10 артефактов",
        "risk": "Пайплайн на словах или уходит на след. год",
        "nextStep": "Неделя 2 — скоринг и классификация сделок",
    },
    "nextId": max([int(d["id"].split("-")[1]) for d in DEALS if str(d["id"]).startswith("D-")] + [0]) + 1,
}

with open(os.path.join(BASE, "js", "initial-data.js"), "w", encoding="utf-8") as f:
    f.write("// Pipeline — auto-generated\nwindow.ITMEN_INITIAL = ")
    json.dump(INITIAL, f, ensure_ascii=False, indent=2)
    f.write(";\n")

print("initial-data.js OK,", len(DEALS), "deals, nextId", INITIAL["nextId"])

# architecture map → architecture-data.js
ARCH_XLSX = os.path.join(ROOT, "Enterprise_IT_Architecture_Map_Export_2026-05-13 (4).xlsx")
if os.path.isfile(ARCH_XLSX):
    wb2 = openpyxl.load_workbook(ARCH_XLSX, data_only=True)
    ws0 = wb2[wb2.sheetnames[0]]
    from collections import OrderedDict
    zones_map = OrderedDict()
    for r in range(2, ws0.max_row + 1):
        zone, cls, vendor, product = ws0.cell(r,1).value, ws0.cell(r,2).value, ws0.cell(r,3).value, ws0.cell(r,4).value
        country, oss = ws0.cell(r,5).value, ws0.cell(r,6).value
        if not zone or not cls: continue
        zones_map.setdefault(zone, OrderedDict()).setdefault(cls, [])
        entry = {"vendor": vendor or "", "product": product or "", "country": country or "", "openSource": oss or ""}
        if entry not in zones_map[zone][cls]:
            zones_map[zone][cls].append(entry)
    ws_cls = wb2["Справочник классов"]
    class_meta = {}
    for r in range(2, ws_cls.max_row + 1):
        name = ws_cls.cell(r,2).value
        if name:
            class_meta[name] = {"id": ws_cls.cell(r,1).value, "what": ws_cls.cell(r,4).value or ""}
    ws_fn = wb2["Функции по классам"]
    functions = {}
    for r in range(2, ws_fn.max_row + 1):
        cls, fn = ws_fn.cell(r,1).value, ws_fn.cell(r,2).value
        if cls and fn:
            functions.setdefault(cls, []).append(fn)
    taxonomy = []
    for zone, classes in zones_map.items():
        cls_list = []
        for cls_name, vendors in classes.items():
            meta = class_meta.get(cls_name, {})
            seen, uniq_fn = set(), []
            for fn in functions.get(cls_name, []):
                if fn not in seen:
                    seen.add(fn)
                    uniq_fn.append(fn)
            cls_list.append({
                "id": meta.get("id") or cls_name[:40],
                "name": cls_name,
                "what": meta.get("what", ""),
                "functions": uniq_fn[:8],
                "catalog": vendors,
            })
        taxonomy.append({"zone": zone, "classes": cls_list})
    arch_out = {"zones": taxonomy, "months": ["Январь","Февраль","Март","Апрель","Май","Июнь","Июль","Август","Сентябрь","Октябрь","Ноябрь","Декабрь"], "years": [2026, 2027, 2028]}
    with open(os.path.join(BASE, "js", "architecture-data.js"), "w", encoding="utf-8") as f:
        f.write("window.ITMEN_ARCHITECTURE = ")
        json.dump(arch_out, f, ensure_ascii=False)
        f.write(";\n")
    print("architecture-data.js OK,", len(taxonomy), "zones")

