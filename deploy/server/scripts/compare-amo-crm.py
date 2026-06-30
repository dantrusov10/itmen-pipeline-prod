#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Compare AmoCRM Excel export vs ITMen Pipeline CRM deals."""
from __future__ import annotations

import json
import os
import re
import sys
import urllib.parse
import urllib.request
from openpyxl import load_workbook

AMO_XLSX = sys.argv[1] if len(sys.argv) > 1 else r"c:\Users\Данила\Downloads\amocrm_export_leads_2026-06-25.xlsx"

LEGAL = re.compile(r"^(ооо|оао|зао|пао|ао|мкпао|ип|фгуп|гбу|гбуз|гбук|гбуко|фгбоу|фгау|фгбу|гку|гуп)\s+", re.I)


def norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = s.replace("«", "").replace("»", "").replace('"', "").replace("'", "")
    s = LEGAL.sub("", s)
    s = s.replace("ё", "е")
    return re.sub(r"\s+", " ", s).strip()


def load_env():
    email = os.environ.get("PB_ADMIN_EMAIL", "")
    password = os.environ.get("PB_ADMIN_PASSWORD", "")
    pb = os.environ.get("PB_URL", "http://127.0.0.1:8095")
    path = "/opt/itmen-pipeline/.env"
    if os.path.exists(path):
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            if k == "PB_ADMIN_EMAIL" and not email:
                email = v
            elif k == "PB_ADMIN_PASSWORD" and not password:
                password = v
            elif k == "PB_URL" and not pb:
                pb = v
    return pb.rstrip("/"), email, password


def http_json(url, data=None, token=None, method=None):
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = token
    body = None if data is None else json.dumps(data).encode("utf-8")
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    with urllib.request.urlopen(req, timeout=120) as res:
        raw = res.read().decode("utf-8")
        return json.loads(raw) if raw else {}


def load_crm_deals():
    local_json = os.environ.get("CRM_DEALS_JSON", "")
    if local_json and os.path.isfile(local_json):
        return json.load(open(local_json, encoding="utf-8"))
    pb, email, password = load_env()
    token = http_json(f"{pb}/api/admins/auth-with-password", {"identity": email, "password": password})["token"]
    deals = []
    page = 1
    while True:
        q = urllib.parse.urlencode({"page": page, "perPage": 200})
        data = http_json(f"{pb}/api/collections/deals/records?{q}", token=token)
        deals.extend(data.get("items", []))
        if page >= data.get("totalPages", 1):
            break
        page += 1
    return deals


def load_amo_leads(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    # First column is always amo ID in standard export
    id_idx = 0
    for i, h in enumerate(header):
        if h.upper() == "ID" or h == "Id":
            id_idx = i
            break
    name_idx = 1
    for i, h in enumerate(header):
        hl = h.lower()
        if "название" in hl and "сдел" in hl:
            name_idx = i
            break
    company_idx = None
    for i, h in enumerate(header):
        hl = h.lower()
        if "компан" in hl and "контакт" not in hl:
            company_idx = i
            break
    if company_idx is None:
        for i, h in enumerate(header):
            if "компан" in h.lower():
                company_idx = i
                break
    leads = []
    for row in rows[1:]:
        if not row or row[id_idx] in (None, ""):
            continue
        try:
            amo_id = int(row[id_idx])
        except (TypeError, ValueError):
            continue
        name = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] is not None else ""
        company = ""
        if company_idx is not None and company_idx < len(row) and row[company_idx] is not None:
            company = str(row[company_idx]).strip()
        leads.append({"amo_id": amo_id, "name": name, "company": company})
    return leads


def main():
    amo = load_amo_leads(AMO_XLSX)
    crm = load_crm_deals()

    amo_ids = {x["amo_id"] for x in amo}
    crm_by_amo = {}
    crm_by_norm = {}
    for d in crm:
        aid = d.get("amoId")
        if aid:
            try:
                crm_by_amo[int(aid)] = d
            except (TypeError, ValueError):
                pass
        cn = norm(d.get("customer") or "")
        if cn and cn not in crm_by_norm:
            crm_by_norm[cn] = d

    missing = []
    for lead in amo:
        aid = lead["amo_id"]
        if aid in crm_by_amo:
            continue
        # fallback: match by normalized company/name
        candidates = [lead["company"], lead["name"]]
        matched = None
        for c in candidates:
            n = norm(c)
            if n and n in crm_by_norm:
                matched = crm_by_norm[n]
                break
        if matched:
            continue
        missing.append(lead)

    print(f"Amo leads: {len(amo)}")
    print(f"CRM deals: {len(crm)} (with amoId: {sum(1 for d in crm if d.get('amoId'))})")
    print(f"Missing in CRM: {len(missing)}")
    print("---")
    for i, m in enumerate(sorted(missing, key=lambda x: x["amo_id"]), 1):
        company = m["company"] or "—"
        name = m["name"] or "—"
        print(f"{i}. amo_id={m['amo_id']} | {name} | компания: {company}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
