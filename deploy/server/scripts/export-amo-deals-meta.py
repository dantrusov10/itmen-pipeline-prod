#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Export AmoCRM deal metadata (fields, contacts, companies, UTM) to Excel."""
from __future__ import annotations

import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone

from openpyxl import Workbook, load_workbook

SUBDOMAIN = os.environ.get("AMO_SUBDOMAIN", "inferit")
BASE = f"https://{SUBDOMAIN}.amocrm.ru"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TOKEN_FILE = os.environ.get("AMO_TOKEN_FILE", os.path.join(SCRIPT_DIR, "amo-tokens.json"))
CLIENT_ID = os.environ.get("AMO_CLIENT_ID", "f4ae4a8e-f973-406a-906a-fc3e29d4a2d9")
CLIENT_SECRET = os.environ.get("AMO_CLIENT_SECRET", "")
REDIRECT_URI = os.environ.get("AMO_REDIRECT_URI", "https://itmen-pipeline.nwlvl.ru/")


def http_json(url, data=None, token=None, method=None):
    headers = {"Content-Type": "application/json", "User-Agent": "ITMen-amo-export/1.0"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    body = None if data is None else json.dumps(data).encode("utf-8")
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=180) as res:
            raw = res.read().decode("utf-8")
            return json.loads(raw) if raw else {}
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {e.code} {url}: {detail}") from e


def get_access_token():
    stored = json.load(open(TOKEN_FILE, encoding="utf-8"))
    if stored.get("refresh_token") and CLIENT_SECRET:
        try:
            data = http_json(
                f"{BASE}/oauth2/access_token",
                {
                    "client_id": CLIENT_ID,
                    "client_secret": CLIENT_SECRET,
                    "grant_type": "refresh_token",
                    "refresh_token": stored["refresh_token"],
                    "redirect_uri": REDIRECT_URI,
                },
            )
            data["saved_at"] = datetime.now(timezone.utc).isoformat()
            json.dump(data, open(TOKEN_FILE, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
            return data["access_token"]
        except Exception:
            pass
    return stored["access_token"]


def ts_fmt(val):
    if not val:
        return ""
    if isinstance(val, int):
        return datetime.fromtimestamp(val, timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    return str(val)


def safe_col(name):
    s = re.sub(r"[\[\]:*?/\\]", " ", str(name or "")).strip()
    return re.sub(r"\s+", " ", s)[:120] or "field"


def load_lead_ids(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    id_idx = idx.get("ID", 0)
    name_idx = idx.get("Название сделки", 1)
    out = {}
    for row in rows[1:]:
        if not row or row[id_idx] in (None, ""):
            continue
        try:
            lid = int(row[id_idx])
        except (TypeError, ValueError):
            continue
        name = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] is not None else ""
        out[lid] = name
    return out


def paginate_embedded(token, path, embedded_key, params=None):
    page = 1
    params = dict(params or {})
    while True:
        q = urllib.parse.urlencode(params)
        url = f"{BASE}{path}?page={page}&limit=250"
        if q:
            url += "&" + q
        data = http_json(url, token=token)
        items = (data.get("_embedded") or {}).get(embedded_key) or []
        for item in items:
            yield item
        if not (data.get("_links") or {}).get("next"):
            break
        page += 1
        time.sleep(0.1)


def fetch_custom_field_map(token, entity):
    fields = {}
    for f in paginate_embedded(token, f"/api/v4/{entity}/custom_fields", "custom_fields"):
        fid = f.get("id")
        if fid is None:
            continue
        fields[int(fid)] = {
            "name": f.get("name") or str(fid),
            "code": f.get("code") or "",
            "type": f.get("type") or "",
        }
    return fields


def fetch_users(token):
    users = {}
    for u in paginate_embedded(token, "/api/v4/users", "users"):
        uid = u.get("id")
        if uid is not None:
            users[int(uid)] = u.get("name") or u.get("login") or str(uid)
    return users


def fetch_pipeline_status_map(token):
    pipelines = {}
    statuses = {}
    data = http_json(f"{BASE}/api/v4/leads/pipelines", token=token)
    for p in (data.get("_embedded") or {}).get("pipelines") or []:
        pid = p.get("id")
        if pid is not None:
            pipelines[int(pid)] = p.get("name") or str(pid)
        for s in (p.get("_embedded") or {}).get("statuses") or []:
            sid = s.get("id")
            if sid is not None:
                statuses[int(sid)] = s.get("name") or str(sid)
    return pipelines, statuses


def cf_values_to_str(values):
    if not values:
        return ""
    parts = []
    for v in values:
        if not isinstance(v, dict):
            parts.append(str(v))
            continue
        val = v.get("value")
        if val is None:
            continue
        extra = v.get("enum_code") or v.get("enum_id")
        if extra and str(extra) not in str(val):
            parts.append(f"{val} ({extra})")
        else:
            parts.append(str(val))
    return "; ".join(parts)


def flatten_custom_fields(custom_fields_values, field_map=None, used_names=None):
    out = {}
    used_names = used_names or set()
    for item in custom_fields_values or []:
        fid = item.get("field_id")
        fname = item.get("field_name")
        if not fname and fid is not None and field_map:
            fname = field_map.get(int(fid), {}).get("name")
        col = safe_col(fname or f"field_{fid}")
        if col in used_names:
            col = safe_col(f"{col}_{fid}")
        used_names.add(col)
        out[col] = cf_values_to_str(item.get("values"))
    return out


def batch_fetch_by_ids(token, entity, ids, with_params=None):
    items = {}
    id_list = sorted(set(ids))
    for i in range(0, len(id_list), 50):
        batch = id_list[i:i + 50]
        params = [("limit", "250")]
        for eid in batch:
            params.append(("filter[id][]", str(eid)))
        if with_params:
            for w in with_params:
                params.append(("with", w))
        q = urllib.parse.urlencode(params)
        url = f"{BASE}/api/v4/{entity}?{q}"
        data = http_json(url, token=token)
        key = entity
        for row in (data.get("_embedded") or {}).get(key) or []:
            rid = row.get("id")
            if rid is not None:
                items[int(rid)] = row
        time.sleep(0.12)
    return items


def fetch_lead_links(token, lead_id):
    links = []
    page = 1
    while True:
        url = f"{BASE}/api/v4/leads/{lead_id}/links?page={page}&limit=250"
        data = http_json(url, token=token)
        links.extend((data.get("_embedded") or {}).get("links") or [])
        if not (data.get("_links") or {}).get("next"):
            break
        page += 1
        time.sleep(0.08)
    return links


def contact_display_name(c):
    parts = [c.get("first_name"), c.get("last_name")]
    name = " ".join(p for p in parts if p).strip()
    return name or c.get("name") or ""


def write_sheet(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((min(len(str(c.value or "")), 60) for c in col[: min(len(col), 150)]), default=10)
        ws.column_dimensions[letter].width = max(10, max_len + 2)


def main():
    xlsx_in = sys.argv[1] if len(sys.argv) > 1 else r"c:\Users\Данила\Downloads\amocrm_export_leads_2026-06-25.xlsx"
    xlsx_out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        os.path.dirname(xlsx_in), "amo_deals_meta_export.xlsx"
    )

    lead_names = load_lead_ids(xlsx_in)
    lead_ids = set(lead_names.keys())
    print(f"leads in file: {len(lead_ids)}")

    token = get_access_token()
    print("token OK")

    users = fetch_users(token)
    pipelines, statuses = fetch_pipeline_status_map(token)
    lead_field_map = fetch_custom_field_map(token, "leads")
    contact_field_map = fetch_custom_field_map(token, "contacts")
    company_field_map = fetch_custom_field_map(token, "companies")
    print(f"fields: leads={len(lead_field_map)} contacts={len(contact_field_map)} companies={len(company_field_map)}")

    leads = batch_fetch_by_ids(token, "leads", lead_ids, with_params=["tags", "loss_reason", "catalog_elements"])
    print(f"leads fetched: {len(leads)}")

    lead_contacts = {}
    lead_companies = {}
    all_contact_ids = set()
    all_company_ids = set()

    for i, lid in enumerate(sorted(lead_ids), 1):
        if lid not in leads:
            continue
        links = fetch_lead_links(token, lid)
        cids = []
        coids = []
        for link in links:
            et = link.get("to_entity_type")
            eid = link.get("to_entity_id")
            if eid is None:
                continue
            try:
                eid = int(eid)
            except (TypeError, ValueError):
                continue
            if et == "contacts":
                cids.append({"id": eid, "is_main": bool(link.get("is_main"))})
                all_contact_ids.add(eid)
            elif et == "companies":
                coids.append({"id": eid, "is_main": bool(link.get("is_main"))})
                all_company_ids.add(eid)
        lead_contacts[lid] = cids
        lead_companies[lid] = coids
        if i % 25 == 0:
            print(f"  links {i}/{len(lead_ids)}")
        time.sleep(0.08)

    contacts = batch_fetch_by_ids(token, "contacts", all_contact_ids)
    companies = batch_fetch_by_ids(token, "companies", all_company_ids)
    print(f"contacts fetched: {len(contacts)}, companies fetched: {len(companies)}")

    deal_rows = []
    deal_headers = set([
        "amo_lead_id", "Название сделки", "Бюджет", "Ответственный", "Воронка", "Этап",
        "Создана", "Обновлена", "Закрыта", "Теги", "Причина отказа", "Создал", "Изменил",
        "Ближайшая задача", "Счёт",
    ])

    for lid in sorted(lead_ids):
        lead = leads.get(lid)
        if not lead:
            deal_rows.append({"amo_lead_id": lid, "Название сделки": lead_names.get(lid, "")})
            continue
        tags = "; ".join(t.get("name", "") for t in (lead.get("_embedded") or {}).get("tags") or [])
        loss = (lead.get("_embedded") or {}).get("loss_reason") or []
        loss_name = loss[0].get("name") if loss else ""
        row = {
            "amo_lead_id": lid,
            "Название сделки": lead.get("name") or lead_names.get(lid, ""),
            "Бюджет": lead.get("price") or 0,
            "Ответственный": users.get(int(lead["responsible_user_id"]), lead.get("responsible_user_id"))
            if lead.get("responsible_user_id") else "",
            "Воронка": pipelines.get(int(lead["pipeline_id"]), lead.get("pipeline_id"))
            if lead.get("pipeline_id") else "",
            "Этап": statuses.get(int(lead["status_id"]), lead.get("status_id"))
            if lead.get("status_id") else "",
            "Создана": ts_fmt(lead.get("created_at")),
            "Обновлена": ts_fmt(lead.get("updated_at")),
            "Закрыта": ts_fmt(lead.get("closed_at")),
            "Теги": tags,
            "Причина отказа": loss_name,
            "Создал": users.get(int(lead["created_by"]), lead.get("created_by")) if lead.get("created_by") else "",
            "Изменил": users.get(int(lead["updated_by"]), lead.get("updated_by")) if lead.get("updated_by") else "",
            "Ближайшая задача": ts_fmt(lead.get("closest_task_at")),
            "Счёт": lead.get("score") or "",
        }
        row.update(flatten_custom_fields(lead.get("custom_fields_values"), lead_field_map))
        deal_rows.append(row)
        deal_headers.update(row.keys())

    contact_rows = []
    contact_headers = set([
        "amo_lead_id", "Название сделки", "contact_id", "Имя", "Основной", "Создан", "Обновлён",
    ])
    for lid in sorted(lead_ids):
        for link in lead_contacts.get(lid, []):
            cid = link["id"]
            c = contacts.get(cid)
            if not c:
                continue
            row = {
                "amo_lead_id": lid,
                "Название сделки": lead_names.get(lid, leads.get(lid, {}).get("name", "")),
                "contact_id": cid,
                "Имя": contact_display_name(c),
                "Основной": "да" if link.get("is_main") else "",
                "Создан": ts_fmt(c.get("created_at")),
                "Обновлён": ts_fmt(c.get("updated_at")),
            }
            row.update(flatten_custom_fields(c.get("custom_fields_values"), contact_field_map))
            contact_rows.append(row)
            contact_headers.update(row.keys())

    company_rows = []
    company_headers = set([
        "amo_lead_id", "Название сделки", "company_id", "Название компании", "Основная", "Создана", "Обновлена",
    ])
    for lid in sorted(lead_ids):
        for link in lead_companies.get(lid, []):
            cid = link["id"]
            co = companies.get(cid)
            if not co:
                continue
            row = {
                "amo_lead_id": lid,
                "Название сделки": lead_names.get(lid, leads.get(lid, {}).get("name", "")),
                "company_id": cid,
                "Название компании": co.get("name") or "",
                "Основная": "да" if link.get("is_main") else "",
                "Создана": ts_fmt(co.get("created_at")),
                "Обновлена": ts_fmt(co.get("updated_at")),
            }
            row.update(flatten_custom_fields(co.get("custom_fields_values"), company_field_map))
            company_rows.append(row)
            company_headers.update(row.keys())

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Сделки"
    deal_header_list = sorted(deal_headers, key=lambda h: (
        0 if h == "amo_lead_id" else 1 if h == "Название сделки" else 2, h
    ))
    write_sheet(ws1, deal_header_list, deal_rows)

    ws2 = wb.create_sheet("Контакты")
    contact_header_list = sorted(contact_headers, key=lambda h: (
        0 if h == "amo_lead_id" else 1 if h == "contact_id" else 2, h
    ))
    write_sheet(ws2, contact_header_list, contact_rows)

    ws3 = wb.create_sheet("Компании")
    company_header_list = sorted(company_headers, key=lambda h: (
        0 if h == "amo_lead_id" else 1 if h == "company_id" else 2, h
    ))
    write_sheet(ws3, company_header_list, company_rows)

    ws4 = wb.create_sheet("Справочник полей")
    catalog = [["Сущность", "field_id", "Код", "Тип", "Название"]]
    for entity, fmap in [("leads", lead_field_map), ("contacts", contact_field_map), ("companies", company_field_map)]:
        for fid, meta in sorted(fmap.items(), key=lambda x: x[1]["name"]):
            catalog.append([entity, fid, meta.get("code"), meta.get("type"), meta.get("name")])
    for row in catalog:
        ws4.append(row)

    wb.save(xlsx_out)
    print(f"saved: {xlsx_out}")
    print(f"deals rows: {len(deal_rows)}, contacts: {len(contact_rows)}, companies: {len(company_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
