#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Migrate Amo data into CRM: amo_id, stages, missing deals, field mapping report."""
from __future__ import annotations

import json
import os
import re
import sys
import urllib.parse
import urllib.request
from collections import Counter
from datetime import datetime, timezone

from openpyxl import load_workbook

PB = "http://127.0.0.1:8095"
LEGAL = re.compile(r"^(ооо|оао|зао|пао|ао|мкпао|ип|фгуп|гбу|гбуз|гбук|гбуко|фгбоу|фгау|фгбу|гку|гуп)\s+", re.I)

FIELD_MAP = {
    "бюджет": ("deals", "amount"),
    "ожидаемый бюджет": ("deals", "expected_budget"),
    "presale": ("deals", "capabilities"),
    "партнер": ("deals", "partner"),
    "партнёр": ("deals", "partner"),
    "конкуренты": ("deals", "competitors"),
    "отрасль": ("deals", "industry"),
    "тип сделки": ("deals", "deal_type"),
    "вероятность": ("deals", "manual_prob"),
    "срок задачи": ("deals", "task_due"),
    "период бюджета": ("deals", "budget_period"),
    "статус бюджета": ("deals", "budget_status"),
    "следующий шаг": ("deals", "next_step_comment"),
    "тип следующего шага": ("deals", "next_step_type"),
    "риск": ("deals", "risk_comment"),
    "тип риска": ("deals", "risk_type"),
    "commit": ("deals", "commit_status"),
    "dml": ("deals", "dml"),
    "utm_source": ("deal_info", "utm_source"),
    "utm_medium": ("deal_info", "utm_medium"),
    "utm_campaign": ("deal_info", "utm_campaign"),
    "utm_content": ("deal_info", "utm_content"),
    "utm_term": ("deal_info", "utm_term"),
    "utm_referrer": ("deal_info", "referrer"),
    "сайт": ("deal_info", "website"),
    "канал": ("deal_info", "source_channel"),
}


def norm(s):
    s = (s or "").strip().lower()
    s = s.replace("«", "").replace("»", "").replace('"', "").replace("'", "")
    s = LEGAL.sub("", s)
    return re.sub(r"\s+", " ", s.replace("ё", "е")).strip()


def tokens(s):
    return set(re.findall(r"[a-zа-я0-9]{4,}", norm(s)))


def match_deal(amo, deal_by_norm, deal_by_amo_id, deals):
    amo_id = amo.get("amo_id")
    if amo_id and amo_id in deal_by_amo_id:
        return deal_by_amo_id[amo_id], "amo_id"

    fields = [
        amo.get("name"),
        amo.get("Компания"),
        amo.get("Компания контакта"),
    ]
    for f in fields:
        n = norm(str(f) if f is not None else "")
        if n and n in deal_by_norm:
            return deal_by_norm[n], "exact"

    amo_toks = set()
    for f in fields:
        amo_toks |= tokens(f)
    if not amo_toks:
        return None, "none"

    best, best_score = None, 0
    for cn, d in deal_by_norm.items():
        ct = tokens(cn)
        if not ct:
            continue
        inter = len(amo_toks & ct)
        if inter >= 2 and inter / min(len(amo_toks), len(ct)) >= 0.5:
            if inter > best_score:
                best_score = inter
                best = d
    if best:
        return best, "fuzzy"
    return None, "none"


def load_env():
    email = password = ""
    for line in open("/opt/itmen-pipeline/.env"):
        if line.startswith("PB_ADMIN_EMAIL="):
            email = line.split("=", 1)[1].strip()
        if line.startswith("PB_ADMIN_PASSWORD="):
            password = line.split("=", 1)[1].strip()
    return email, password


def pb_token(email, password):
    return json.loads(urllib.request.urlopen(urllib.request.Request(
        PB + "/api/admins/auth-with-password",
        json.dumps({"identity": email, "password": password}).encode(),
        headers={"Content-Type": "application/json"},
    )).read())["token"]


def pb_list(token, collection, filter_q="", sort=""):
    items, page = [], 1
    while True:
        q = urllib.parse.urlencode({"page": page, "perPage": 200, "filter": filter_q, "sort": sort})
        data = json.loads(urllib.request.urlopen(urllib.request.Request(
            f"{PB}/api/collections/{collection}/records?{q}", headers={"Authorization": token},
        )).read())
        items.extend(data.get("items", []))
        if page >= data.get("totalPages", 1):
            break
        page += 1
    return items


def pb_patch(token, collection, rec_id, body):
    clean = {k: v for k, v in body.items() if v is not None}
    req = urllib.request.Request(
        f"{PB}/api/collections/{collection}/records/{rec_id}",
        json.dumps(clean).encode(),
        headers={"Authorization": token, "Content-Type": "application/json"},
        method="PATCH",
    )
    return json.loads(urllib.request.urlopen(req).read())


def pb_create(token, collection, body):
    req = urllib.request.Request(
        f"{PB}/api/collections/{collection}/records",
        json.dumps(body).encode(),
        headers={"Authorization": token, "Content-Type": "application/json"},
        method="POST",
    )
    return json.loads(urllib.request.urlopen(req).read())


def load_amo_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    out = []
    for row in rows[1:]:
        if not row or row[idx.get("ID", 0)] in (None, ""):
            continue
        rec = {"amo_id": int(row[idx["ID"]]), "name": str(row[idx.get("Название сделки", 1)] or "").strip()}
        for h in header:
            if h in ("ID",):
                continue
            v = row[idx[h]] if idx[h] < len(row) else None
            if v is not None and str(v).strip() != "":
                rec[h] = v
        out.append(rec)
    return out


def load_meta_deals(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Сделки"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    out = {}
    for row in rows[1:]:
        if not row:
            continue
        lid = row[idx.get("amo_lead_id", 0)]
        if lid in (None, ""):
            continue
        rec = {}
        for h in header:
            v = row[idx[h]] if idx[h] < len(row) else None
            rec[h] = v
        out[int(lid)] = rec
    return out


def next_deal_id(deals, meta_rows):
    mx = 0
    for d in deals:
        m = re.match(r"D-(\d+)", d.get("deal_id") or "")
        if m:
            mx = max(mx, int(m.group(1)))
    for m in meta_rows:
        if m.get("id"):
            mm = re.match(r"D-(\d+)", str(m["id"]))
            if mm:
                mx = max(mx, int(mm.group(1)))
    return mx + 1


def map_field_name(name):
    n = norm(name)
    if n in FIELD_MAP:
        return FIELD_MAP[n]
    for key, target in FIELD_MAP.items():
        if key in n:
            return target
    return None


def deal_amo_id(deal):
    for key in ("amo_id", "amoId"):
        v = deal.get(key)
        if v not in (None, "", 0):
            try:
                return int(v)
            except (TypeError, ValueError):
                pass
    return None


def main():
    amo_xlsx = sys.argv[1] if len(sys.argv) > 1 else "/tmp/amocrm_export_leads_2026-06-25.xlsx"
    meta_xlsx = sys.argv[2] if len(sys.argv) > 2 else "/tmp/amo_deals_meta_export.xlsx"
    report_path = sys.argv[3] if len(sys.argv) > 3 else "/tmp/amo-crm-migration-report.json"
    apply = "--apply" in sys.argv

    amo_rows = load_amo_xlsx(amo_xlsx)
    meta_by_id = load_meta_deals(meta_xlsx) if os.path.isfile(meta_xlsx) else {}

    email, password = load_env()
    token = pb_token(email, password)
    deals = pb_list(token, "deals")
    stages = [x["value"] for x in pb_list(token, "list_items", 'list_key="stages"')]
    stage_norm = {norm(s): s for s in stages}

    deal_by_norm = {}
    deal_by_amo_id = {}
    for d in deals:
        cn = norm(d.get("customer") or "")
        if cn and cn not in deal_by_norm:
            deal_by_norm[cn] = d
        aid = deal_amo_id(d)
        if aid not in (None, "", 0):
            try:
                deal_by_amo_id[int(aid)] = d
            except (TypeError, ValueError):
                pass

    report = {
        "amo_id_updated": [],
        "amo_id_skipped": [],
        "stage_updated": [],
        "stage_uncertain": [],
        "missing_created": [],
        "field_mapping": {"mapped": [], "uncertain": [], "unused_amo_fields": []},
        "unmatched_amo": [],
        "match_stats": {"amo_id": 0, "exact": 0, "fuzzy": 0},
    }

    used_deal_ids = set()
    for amo in amo_rows:
        deal, match_kind = match_deal(amo, deal_by_norm, deal_by_amo_id, deals)
        if not deal:
            report["unmatched_amo"].append({
                "amo_id": amo["amo_id"],
                "name": amo["name"],
                "company": str(amo.get("Компания") or ""),
                "contact_company": str(amo.get("Компания контакта") or ""),
            })
            continue
        if deal["id"] in used_deal_ids:
            report["amo_id_skipped"].append({
                "deal_id": deal["deal_id"],
                "customer": deal["customer"],
                "amo_id": amo["amo_id"],
                "reason": "duplicate_crm_deal",
            })
            continue
        used_deal_ids.add(deal["id"])
        report["match_stats"][match_kind] = report["match_stats"].get(match_kind, 0) + 1
        patch = {"amo_id": amo["amo_id"]}
        if apply and deal_amo_id(deal) != amo["amo_id"]:
            pb_patch(token, "deals", deal["id"], patch)
        report["amo_id_updated"].append({
            "deal_id": deal["deal_id"],
            "customer": deal["customer"],
            "amo_id": amo["amo_id"],
            "match": match_kind,
        })

        meta = meta_by_id.get(amo["amo_id"], {})
        amo_stage = str(meta.get("Этап") or amo.get("Этап сделки") or "").strip()
        if amo_stage:
            sn = norm(amo_stage)
            crm_stage = stage_norm.get(sn)
            if crm_stage and crm_stage != deal.get("stage"):
                report["stage_updated"].append({
                    "deal_id": deal["deal_id"], "from": deal.get("stage"), "to": crm_stage, "amo_stage": amo_stage,
                })
                if apply:
                    pb_patch(token, "deals", deal["id"], {"stage": crm_stage})
            elif not crm_stage:
                report["stage_uncertain"].append({
                    "deal_id": deal["deal_id"], "amo_stage": amo_stage, "crm_stages_sample": stages[:15],
                })

    # create missing
    meta = pb_list(token, "pipeline_meta", 'slug="main"')
    next_id = next_deal_id(deals, [])
    for miss in report["unmatched_amo"]:
        amo = next(a for a in amo_rows if a["amo_id"] == miss["amo_id"])
        deal_id = f"D-{next_id:03d}"
        next_id += 1
        amo_stage = str(amo.get("Этап сделки") or meta_by_id.get(amo["amo_id"], {}).get("Этап") or "Новая").strip()
        crm_stage = stage_norm.get(norm(amo_stage), stages[0] if stages else "Новая")
        body = {
            "deal_id": deal_id,
            "customer": amo["name"],
            "owner": str(amo.get("Ответственный") or ""),
            "stage": crm_stage,
            "amount": float(amo.get("Бюджет") or 0) if amo.get("Бюджет") not in (None, "") else 0,
            "amo_id": amo["amo_id"],
            "lastUpdate": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        }
        if apply:
            created = pb_create(token, "deals", body)
            report["missing_created"].append({"deal_id": deal_id, "customer": amo["name"], "amo_id": amo["amo_id"], "pb_id": created["id"]})
        else:
            report["missing_created"].append({"deal_id": deal_id, "customer": amo["name"], "amo_id": amo["amo_id"], "would_create": True})

    # custom fields analysis from meta export
    if meta_by_id:
        col_nonempty = Counter()
        for rec in meta_by_id.values():
            for k, v in rec.items():
                if k in ("amo_lead_id", "Название сделки"):
                    continue
                if v is not None and str(v).strip() not in ("", "None"):
                    col_nonempty[k] += 1
        skip_std = {"Бюджет", "Ответственный", "Воронка", "Этап", "Создана", "Обновлена", "Закрыта", "Теги", "Причина отказа", "Создал", "Изменил", "Ближайшая задача", "Счёт"}
        for col, cnt in col_nonempty.most_common():
            if col in skip_std:
                continue
            target = map_field_name(col)
            if target:
                report["field_mapping"]["mapped"].append({"amo_field": col, "filled_rows": cnt, "target": f"{target[0]}.{target[1]}"})
            else:
                report["field_mapping"]["uncertain"].append({"amo_field": col, "filled_rows": cnt})

    if meta:
        if apply and report["missing_created"]:
            pb_patch(token, "pipeline_meta", meta[0]["id"], {"next_id": next_id})

    mode = "APPLY" if apply else "DRY-RUN"
    print(f"migrate-amo-crm [{mode}]")
    print(json.dumps({
        "amo_id": len(report["amo_id_updated"]),
        "match_stats": report["match_stats"],
        "stage_updates": len(report["stage_updated"]),
        "stage_uncertain": len(report["stage_uncertain"]),
        "missing": len(report["unmatched_amo"]),
        "missing_created": len(report["missing_created"]),
        "fields_mapped": len(report["field_mapping"]["mapped"]),
        "fields_uncertain": len(report["field_mapping"]["uncertain"]),
    }, ensure_ascii=False, indent=2))
    json.dump(report, open(report_path, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(f"report: {report_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
