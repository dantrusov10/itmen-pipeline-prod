#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import json
import re
import sys
from openpyxl import load_workbook

AMO = sys.argv[1]
CRM = sys.argv[2]

LEGAL = re.compile(
    r"^(ооо|оао|зао|пао|ао|мкпао|ип|фгуп|гбу|гбуз|гбук|гбуко|фгбоу|фгау|фгбу|гку|гуп)\s+",
    re.I,
)


def norm(s):
    s = (s or "").strip().lower()
    s = s.replace("«", "").replace("»", "").replace('"', "").replace("'", "")
    s = LEGAL.sub("", s)
    return re.sub(r"\s+", " ", s.replace("ё", "е")).strip()


def tokens(s):
    return set(re.findall(r"[a-zа-я0-9]{4,}", norm(s)))


crm = json.load(open(CRM, encoding="utf-8"))
crm_norm = {}
for d in crm:
    n = norm(d.get("customer"))
    if n:
        crm_norm[n] = d

wb = load_workbook(AMO, read_only=True, data_only=True)
rows = list(wb.active.iter_rows(values_only=True))
header = [str(x).strip() if x is not None else "" for x in rows[0]]
idx = {h: i for i, h in enumerate(header)}

missing = []
matched = 0
for row in rows[1:]:
    if not row or row[0] in (None, ""):
        continue
    amo_id = int(row[0])
    name = row[idx["Название сделки"]] if "Название сделки" in idx else row[1]
    company = row[idx["Компания"]] if "Компания" in idx else None
    contact_co = row[idx["Компания контакта"]] if "Компания контакта" in idx else None
    stage = row[idx["Этап сделки"]] if "Этап сделки" in idx else ""
    owner = row[idx["Ответственный"]] if "Ответственный" in idx else ""

    fields = [name, company, contact_co]
    hit = None
    for f in fields:
        n = norm(str(f) if f is not None else "")
        if n and n in crm_norm:
            hit = crm_norm[n]
            break
    if hit:
        matched += 1
        continue

    # soft: significant token overlap
    best = None
    best_score = 0
    amo_toks = set()
    for f in fields:
        amo_toks |= tokens(f)
    if amo_toks:
        for cn, d in crm_norm.items():
            ct = tokens(cn)
            if not ct:
                continue
            inter = len(amo_toks & ct)
            if inter >= 2 and inter / min(len(amo_toks), len(ct)) >= 0.5:
                score = inter
                if score > best_score:
                    best_score = score
                    best = d
    if best:
        matched += 1
        continue

    missing.append({
        "amo_id": amo_id,
        "name": str(name or "").strip(),
        "company": str(company or "").strip(),
        "contact_company": str(contact_co or "").strip(),
        "stage": str(stage or "").strip(),
        "owner": str(owner or "").strip(),
    })

sys.stdout.reconfigure(encoding="utf-8")
print(f"Amo: {len(rows)-1} | CRM: {len(crm)} | matched: {matched} | missing: {len(missing)}")
print("---")
for i, m in enumerate(missing, 1):
    print(f"{i}. amo_id={m['amo_id']}")
    print(f"   Название: {m['name'] or '—'}")
    print(f"   Компания: {m['company'] or '—'}")
    if m["contact_company"]:
        print(f"   Компания контакта: {m['contact_company']}")
    print(f"   Этап: {m['stage'] or '—'} | Ответственный: {m['owner'] or '—'}")
