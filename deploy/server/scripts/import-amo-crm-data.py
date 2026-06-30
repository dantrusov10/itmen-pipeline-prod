#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Массовый импорт данных AmoCRM в PocketBase (CRM v4).

  python3 import-amo-crm-data.py                         # dry-run
  python3 import-amo-crm-data.py --apply                   # импорт
  python3 import-amo-crm-data.py --apply --skip-files      # без вложений

Файлы по умолчанию в /tmp/ (загрузить с Downloads перед запуском).
"""
from __future__ import annotations

import argparse
import json
import mimetypes
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from collections import defaultdict
from datetime import datetime, timezone
from uuid import uuid4
from io import BytesIO

from openpyxl import load_workbook

PB = os.environ.get("PB_URL", "http://127.0.0.1:8095").rstrip("/")

FIELD_MAP = {
    "бюджет": ("deals", "amount"),
    "ожидаемый бюджет": ("deals", "expected_budget"),
    "presale": ("deals", "capabilities"),
    "партнер": ("deals", "partner"),
    "партнёр": ("deals", "partner"),
    "конкуренты": ("deals", "competitors"),
    "отрасль": ("deals", "industry"),
    "тип сделки": ("deals", "deal_type"),
    "вероятность": ("deals", "manual_prob"),
    "шанс закрытия": ("deals", "manual_prob"),
    "срок задачи": ("deals", "task_due"),
    "период бюджета": ("deals", "budget_period"),
    "статус бюджета": ("deals", "budget_status"),
    "следующий шаг": ("deals", "next_step_comment"),
    "тип следующего шага": ("deals", "next_step_type"),
    "риск": ("deals", "risk_comment"),
    "тип риска": ("deals", "risk_type"),
    "commit": ("deals", "commit_status"),
    "dml": ("deals", "dml"),
    "маржа партнера": ("deals", "partner_discount"),
    "процент скидки клиенту": ("deals", "client_discount"),
    "utm_source": ("deal_info", "utm_source"),
    "utm_medium": ("deal_info", "utm_medium"),
    "utm_campaign": ("deal_info", "utm_campaign"),
    "utm_content": ("deal_info", "utm_content"),
    "utm_term": ("deal_info", "utm_term"),
    "utm_referrer": ("deal_info", "referrer"),
    "сайт": ("deal_info", "website"),
    "канал привлечения": ("deal_info", "source_channel"),
    "канал продаж": ("deal_info", "source_channel"),
    "дата привлечения": ("deal_info", "lead_date"),
    "оборот сделки": ("deals", "expected_budget"),
    "сумма сделки": ("deals", "amount"),
}

AMO_INFO_EXACT = {
    "продукт итмен": "product_itmen",
    "конечные точки": "endpoints",
    "формат закупки": "procurement_format",
    "регистрация до даты": "registration_deadline",
    "общий размер инфраструктуры": "infrastructure_size",
    "грейд": "grade",
    "инструмент закрытия": "closing_tool",
    "соответствие функицоналу": "functional_fit",
    "соответствие функционалу": "functional_fit",
    "старт теста": "test_start",
    "окончание теста": "test_end",
    "дистрибьютор": "distributor",
    "вид деятельности": "activity_kind",
    "ос по тестированию": "test_os",
    "планируемая дата оплаты": "planned_payment_date",
    "дата отгрузки": "shipment_date",
    "ссылка на карту проекта": "project_map_url",
    "abm_tier": "abm_tier",
    "срок контракта": "contract_term",
}

SKIP_AMO_FIELDS = {
    "бизнес задачи", "ссылка на kaiten", "product_focus", "product focus",
}

INFO_DATE_FIELDS = {
    "lead_date", "registration_deadline", "test_start", "test_end",
    "planned_payment_date", "shipment_date",
}

SKIP_DEAL_COLS = {
    "amo_lead_id", "Название сделки", "Бюджет", "Ответственный", "Воронка", "Этап",
    "Создана", "Обновлена", "Закрыта", "Теги", "Причина отказа", "Создал", "Изменил",
    "Ближайшая задача", "Счёт",
}

EMAIL_COLS = ("email", "e-mail", "рабочий email", "email рабочий", "личный email")
PHONE_COLS = ("телефон", "phone", "рабочий телефон", "мобильный телефон", "рабочий", "мобильный")
ROLE_COLS = ("должность", "роль", "position")

ACTIVITY_TYPE_MAP = {
    "комментарий": "comment",
    "входящий звонок": "call_in",
    "исходящий звонок": "call_out",
    "входящее sms": "sms_in",
    "исходящее sms": "sms_out",
    "системное": "system",
    "системное (рас.)": "system",
    "файл": "file_note",
    "геолокация": "geolocation",
    "задача (результат)": "task_done",
}


def norm(s):
    return re.sub(r"\s+", " ", (s or "").strip().lower().replace("ё", "е"))


def map_field_name(name):
    n = norm(name)
    if n in SKIP_AMO_FIELDS:
        return None
    if n in AMO_INFO_EXACT:
        return ("deal_info", AMO_INFO_EXACT[n])
    if n in FIELD_MAP:
        return FIELD_MAP[n]
    for key, target in AMO_INFO_EXACT.items():
        if key in n:
            return ("deal_info", target)
    for key, target in FIELD_MAP.items():
        if key in n:
            return target
    return None


def load_env():
    email = password = ""
    pb = PB
    for path in ("/opt/itmen-pipeline/.env",):
        if not os.path.isfile(path):
            continue
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            if k == "PB_ADMIN_EMAIL":
                email = v
            elif k == "PB_ADMIN_PASSWORD":
                password = v
            elif k == "PB_URL":
                pb = v.rstrip("/")
    return pb, email, password


def http_json(url, data=None, token=None, method=None, timeout=120):
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = token
    body = None if data is None else json.dumps(data, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as res:
            raw = res.read().decode("utf-8")
            return json.loads(raw) if raw else {}
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {e.code} {url}: {detail}") from e


def pb_token(pb, email, password):
    return http_json(f"{pb}/api/admins/auth-with-password",
                     {"identity": email, "password": password})["token"]


def pb_list(token, collection, filter_q="", fields=""):
    items, page = [], 1
    while True:
        params = {"page": page, "perPage": 200}
        if filter_q:
            params["filter"] = filter_q
        if fields:
            params["fields"] = fields
        q = urllib.parse.urlencode(params)
        data = http_json(f"{PB}/api/collections/{collection}/records?{q}", token=token)
        items.extend(data.get("items", []))
        if page >= data.get("totalPages", 1):
            break
        page += 1
    return items


def pb_create(token, collection, body):
    return http_json(f"{PB}/api/collections/{collection}/records", body, token=token, method="POST")


def pb_patch(token, collection, rec_id, body):
    clean = {k: v for k, v in body.items() if v is not None}
    return http_json(f"{PB}/api/collections/{collection}/records/{rec_id}", clean,
                     token=token, method="PATCH")


def pb_upload_file(token, collection, fields: dict, file_bytes: bytes, filename: str, file_field="file"):
    boundary = uuid4().hex
    body = BytesIO()
    for key, val in fields.items():
        if val is None:
            continue
        body.write(f"--{boundary}\r\n".encode())
        body.write(f'Content-Disposition: form-data; name="{key}"\r\n\r\n'.encode())
        body.write(f"{val}\r\n".encode("utf-8"))
    mime = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    safe_name = filename.replace('"', "_")[:280]
    body.write(f"--{boundary}\r\n".encode())
    body.write(
        f'Content-Disposition: form-data; name="{file_field}"; filename="{safe_name}"\r\n'
        f"Content-Type: {mime}\r\n\r\n".encode()
    )
    body.write(file_bytes)
    body.write(f"\r\n--{boundary}--\r\n".encode())
    payload = body.getvalue()
    req = urllib.request.Request(
        f"{PB}/api/collections/{collection}/records",
        data=payload,
        headers={"Authorization": token, "Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=300) as res:
            return json.loads(res.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {e.code} upload {filename}: {detail}") from e


def read_xlsx_sheet(path, sheet_name=None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    out = []
    for row in rows[1:]:
        rec = {}
        for i, h in enumerate(header):
            if not h:
                continue
            v = row[i] if i < len(row) else None
            rec[h] = v
        out.append(rec)
    return header, out


def parse_dt(val):
    if val in (None, ""):
        return None
    if isinstance(val, datetime):
        return val.replace(tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d", "%d.%m.%Y %H:%M:%S", "%d.%m.%Y"):
        try:
            dt = datetime.strptime(s[:19] if fmt.endswith("%S") else s[:10], fmt)
            return dt.replace(tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")
        except ValueError:
            pass
    return None


def deal_amo_id(deal):
    for key in ("amo_id", "amoId"):
        v = deal.get(key)
        if v not in (None, "", 0):
            try:
                return int(v)
            except (TypeError, ValueError):
                pass
    return None


def build_amo_deal_map(deals):
    m = {}
    for d in deals:
        aid = deal_amo_id(d)
        if aid is not None:
            m[aid] = d
    return m


def pick_col(header_map, candidates):
    for c in candidates:
        for h, n in header_map.items():
            if norm(h) == norm(c) or norm(c) in norm(h):
                return h
    return None


def contact_fields_from_row(row):
    header_map = {h: norm(h) for h in row.keys()}
    email = phone = role = ""
    for h, v in row.items():
        nh = norm(h)
        if not email and any(k in nh for k in EMAIL_COLS):
            email = str(v or "").strip()
        if not phone and any(k in nh for k in PHONE_COLS):
            phone = str(v or "").strip()
        if not role and any(k in nh for k in ROLE_COLS):
            role = str(v or "").strip()
    return email, phone, role


def company_inn_from_row(row):
    for h, v in row.items():
        if "инн" in norm(h):
            return str(v or "").strip()
    return ""


def coerce_field_value(target_col, val):
    if val in (None, ""):
        return None
    if target_col in ("amount", "expected_budget", "manual_prob", "partner_discount", "client_discount"):
        try:
            return float(str(val).replace(",", ".").replace("%", "").replace(" ", "").strip())
        except ValueError:
            return None
    if target_col in INFO_DATE_FIELDS:
        return parse_dt(val)
    if target_col == "task_due":
        s = str(val).strip()
        return s[:10] if s else None
    return str(val).strip()


def ensure_v4_collections(apply):
    script = os.path.join(os.path.dirname(__file__), "apply-crm-v4-collections.py")
    if not os.path.isfile(script):
        print("  apply-crm-v4-collections.py not found — skip schema check")
        return
    import subprocess
    cmd = [sys.executable, script]
    if apply:
        cmd.append("--apply")
    subprocess.run(cmd, check=False)


def import_deal_fields(meta_deals, amo_map, deal_info_existing, apply, overwrite_empty, report):
    deals_patch = defaultdict(dict)
    info_patch = defaultdict(dict)

    for amo_id, row in meta_deals.items():
        deal = amo_map.get(amo_id)
        if not deal:
            report["skipped_no_deal"] += 1
            continue
        pb_id = deal["id"]

        loss = str(row.get("Причина отказа") or "").strip()
        if loss and (overwrite_empty or not deal.get("loss_reason")):
            deals_patch[pb_id]["loss_reason"] = loss

        for col, val in row.items():
            if col in SKIP_DEAL_COLS or val in (None, ""):
                continue
            if norm(col) in SKIP_AMO_FIELDS:
                continue
            target = map_field_name(col)
            if not target:
                continue
            coll, field = target
            coerced = coerce_field_value(field, val)
            if coerced is None:
                continue
            if coll == "deals":
                if not overwrite_empty and deal.get(field) not in (None, "", 0):
                    continue
                deals_patch[pb_id][field] = coerced
            else:
                existing = deal_info_existing.get(pb_id, {})
                if not overwrite_empty and existing.get(field) not in (None, ""):
                    continue
                info_patch[pb_id][field] = coerced

        created = parse_dt(row.get("Создана"))
        if created and pb_id in info_patch:
            if overwrite_empty or not deal_info_existing.get(pb_id, {}).get("lead_date"):
                info_patch[pb_id].setdefault("lead_date", created)
        if str(row.get("Бюджет") or "").strip() and (overwrite_empty or not deal.get("amount")):
            b = coerce_field_value("amount", row.get("Бюджет"))
            if b is not None:
                deals_patch[pb_id]["amount"] = b

    report["deals_fields_patched"] = len(deals_patch)
    report["deal_info_upserted"] = len(info_patch)

    if not apply:
        return

    for pb_id, patch in deals_patch.items():
        pb_patch(token_global, "deals", pb_id, patch)

    for pb_id, patch in info_patch.items():
        patch["deal"] = pb_id
        existing = deal_info_existing.get(pb_id)
        if existing:
            pb_patch(token_global, "deal_info", existing["id"], patch)
        else:
            pb_create(token_global, "deal_info", patch)


def trunc(s, n):
    s = str(s or "").strip()
    return s[:n] if len(s) > n else s


def import_contacts(contacts_rows, amo_map, existing_contacts, apply, force, report):
    by_deal = defaultdict(list)
    for row in contacts_rows:
        try:
            amo_id = int(row.get("amo_lead_id"))
        except (TypeError, ValueError):
            continue
        deal = amo_map.get(amo_id)
        if not deal:
            continue
        name = str(row.get("Имя") or "").strip()
        if not name:
            continue
        email, phone, role = contact_fields_from_row(row)
        is_primary = str(row.get("Основной") or "").strip().lower() in ("да", "yes", "1", "true")
        by_deal[deal["id"]].append({
            "name": trunc(name, 200),
            "email": trunc(email, 200),
            "phone": trunc(phone, 80),
            "role": trunc(role, 120),
            "is_primary": is_primary,
        })

    report["contacts_deals"] = len(by_deal)
    report["contacts_rows"] = sum(len(v) for v in by_deal.values())

    if not apply:
        return

    for pb_id, contacts in by_deal.items():
        if existing_contacts.get(pb_id) and not force:
            report["contacts_skipped_existing"] += 1
            continue
        for c in existing_contacts.get(pb_id, []):
            req = urllib.request.Request(
                f"{PB}/api/collections/deal_contacts/records/{c['id']}",
                headers={"Authorization": token_global}, method="DELETE")
            urllib.request.urlopen(req, timeout=60)
        for i, c in enumerate(contacts):
            pb_create(token_global, "deal_contacts", {
                "deal": pb_id,
                "name": c["name"],
                "email": c["email"],
                "phone": c["phone"],
                "role": c["role"],
                "sort_order": i,
                "is_primary": c["is_primary"],
            })
        report["contacts_imported_deals"] += 1


def import_companies(companies_rows, amo_map, deal_info_existing, apply, overwrite_empty, report):
    by_deal = {}
    for row in companies_rows:
        try:
            amo_id = int(row.get("amo_lead_id"))
        except (TypeError, ValueError):
            continue
        if str(row.get("Основная") or "").strip().lower() not in ("да", "yes", "1", "true", ""):
            if amo_id in by_deal:
                continue
        deal = amo_map.get(amo_id)
        if not deal:
            continue
        name = str(row.get("Название компании") or "").strip()
        inn = company_inn_from_row(row)
        if name or inn:
            by_deal[deal["id"]] = {"company_name": name, "company_inn": inn}

    report["companies_deals"] = len(by_deal)
    if not apply:
        return

    for pb_id, patch in by_deal.items():
        existing = deal_info_existing.get(pb_id, {})
        body = {"deal": pb_id}
        if patch.get("company_name") and (overwrite_empty or not existing.get("company_name")):
            body["company_name"] = patch["company_name"]
        if patch.get("company_inn") and (overwrite_empty or not existing.get("company_inn")):
            body["company_inn"] = patch["company_inn"]
        if len(body) <= 1:
            continue
        if existing:
            pb_patch(token_global, "deal_info", existing["id"], body)
        else:
            pb_create(token_global, "deal_info", body)


def activity_type(label):
    return ACTIVITY_TYPE_MAP.get(norm(label), "comment")


def import_activities(comments_rows, stage_rows, amo_map, existing_refs, apply, report):
    to_create = []

    for row in comments_rows:
        try:
            amo_id = int(row.get("amo_lead_id"))
            source_id = row.get("source_id")
        except (TypeError, ValueError):
            continue
        deal = amo_map.get(amo_id)
        if not deal or source_id in (None, ""):
            continue
        ref = f"amo:note:{source_id}"
        if ref in existing_refs:
            report["activities_skipped_dup"] += 1
            continue
        label = str(row.get("Тип") or row.get("event_type") or "Комментарий")
        body = str(row.get("Текст") or "").strip()
        task_txt = str(row.get("Задача") or row.get("task_text") or "").strip()
        if task_txt and label.lower().startswith("задача"):
            body = f"{task_txt}\n\n{body}".strip()
        meta = {"amo_type": label, "amo_lead_id": amo_id}
        to_create.append({
            "deal": deal["id"],
            "activity_type": activity_type(label),
            "body": body or label,
            "author": trunc(row.get("Автор") or row.get("author") or "", 120),
            "meta_json": json.dumps(meta, ensure_ascii=False),
            "activity_at": parse_dt(row.get("Дата") or row.get("created_at")) or datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "ref_id": ref,
        })

    for row in stage_rows:
        try:
            amo_id = int(row.get("amo_lead_id"))
            event_id = row.get("event_id")
        except (TypeError, ValueError):
            continue
        deal = amo_map.get(amo_id)
        if not deal or event_id in (None, ""):
            continue
        ref = f"amo:stage:{event_id}"
        if ref in existing_refs:
            report["activities_skipped_dup"] += 1
            continue
        before = str(row.get("Было") or "")
        after = str(row.get("Стало") or "")
        to_create.append({
            "deal": deal["id"],
            "activity_type": "stage_change",
            "body": f"{before} → {after}".strip(" →"),
            "author": trunc(row.get("Автор") or "", 120),
            "meta_json": json.dumps({
                "from": before, "to": after, "pipeline": str(row.get("Воронка") or ""),
                "amo_lead_id": amo_id,
            }, ensure_ascii=False),
            "activity_at": parse_dt(row.get("Дата")) or datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "ref_id": ref,
        })

    report["activities_to_create"] = len(to_create)
    if not apply:
        return

    batch = 0
    for rec in to_create:
        pb_create(token_global, "deal_activities", rec)
        existing_refs.add(rec["ref_id"])
        batch += 1
        if batch % 100 == 0:
            print(f"  activities {batch}/{len(to_create)}")


def import_tasks(open_rows, comments_rows, amo_map, existing_task_keys, existing_amo_task_ids, apply, report):
    to_create = []

    for row in open_rows:
        try:
            amo_id = int(row.get("amo_lead_id"))
            task_id = row.get("task_id")
        except (TypeError, ValueError):
            continue
        deal = amo_map.get(amo_id)
        if not deal or task_id in (None, ""):
            continue
        if str(task_id) in existing_amo_task_ids:
            report["tasks_skipped_dup"] += 1
            continue
        title = str(row.get("Задача") or "").strip()
        if not title:
            continue
        due = parse_dt(row.get("Срок"))
        key = (deal["id"], title, (due or "")[:19])
        if key in existing_task_keys:
            report["tasks_skipped_dup"] += 1
            continue
        to_create.append({
            "deal": deal["id"],
            "title": title[:300],
            "assignee": trunc(row.get("Ответственный") or "", 120),
            "due_at": due,
            "status": "open",
            "created_by": "Amo import",
            "description": f"amo_task_id={task_id}",
        })
        existing_task_keys.add(key)
        existing_amo_task_ids.add(str(task_id))

    for row in comments_rows:
        label = norm(str(row.get("Тип") or row.get("event_type") or ""))
        if "задача" not in label or "результат" not in label:
            continue
        try:
            amo_id = int(row.get("amo_lead_id"))
            source_id = row.get("source_id")
        except (TypeError, ValueError):
            continue
        deal = amo_map.get(amo_id)
        if not deal:
            continue
        ref_done = f"amo_task_done:{source_id}"
        if source_id and ref_done in existing_amo_task_ids:
            report["tasks_skipped_dup"] += 1
            continue
        title = str(row.get("Задача") or row.get("task_text") or "Задача").strip()
        body = str(row.get("Текст") or "").strip()
        done_at = parse_dt(row.get("Дата") or row.get("created_at"))
        key = (deal["id"], title, (done_at or "")[:19])
        if key in existing_task_keys:
            report["tasks_skipped_dup"] += 1
            continue
        to_create.append({
            "deal": deal["id"],
            "title": title[:300],
            "description": f"amo_task_done:{source_id}\n{body}".strip(),
            "assignee": trunc(row.get("Автор") or row.get("author") or "", 120),
            "due_at": done_at,
            "done_at": done_at,
            "status": "done",
            "created_by": "Amo import",
        })
        existing_task_keys.add(key)
        if source_id:
            existing_amo_task_ids.add(ref_done)

    report["tasks_to_create"] = len(to_create)
    if not apply:
        return

    batch = 0
    for rec in to_create:
        pb_create(token_global, "deal_tasks", rec)
        batch += 1
        if batch % 50 == 0:
            print(f"  tasks {batch}/{len(to_create)}")


def import_files(manifest_rows, zip_path, amo_map, existing_file_labels, apply, report):
    if not os.path.isfile(zip_path):
        report["files_error"] = f"zip not found: {zip_path}"
        return

    ok_rows = [r for r in manifest_rows if str(r.get("download_status") or "").startswith("ok")]
    report["files_manifest_ok"] = len(ok_rows)

    with zipfile.ZipFile(zip_path, "r") as zf:
        names = set(zf.namelist())
        batch = 0
        for row in ok_rows:
            try:
                amo_id = int(row.get("amo_lead_id"))
            except (TypeError, ValueError):
                continue
            deal = amo_map.get(amo_id)
            if not deal:
                continue
            arc = str(row.get("Архивный путь") or "")
            if arc not in names:
                report["files_missing_in_zip"] += 1
                continue
            note_id = row.get("note_id")
            label = f"Amo:{note_id}"
            if label in existing_file_labels.get(deal["id"], set()):
                report["files_skipped_dup"] += 1
                continue
            fname = str(row.get("Имя файла") or os.path.basename(arc))
            if not apply:
                report["files_to_upload"] += 1
                continue
            try:
                content = zf.read(arc)
                uploaded_at = parse_dt(row.get("Дата")) or datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
                pb_upload_file(token_global, "deal_files", {
                    "deal": deal["id"],
                    "label": trunc(label, 80),
                    "original_name": trunc(fname, 300),
                    "size": len(content),
                    "mime_type": trunc(mimetypes.guess_type(fname)[0] or "application/octet-stream", 120),
                    "uploaded_by": trunc(str(row.get("Автор") or "Amo import"), 120),
                    "uploaded_at": uploaded_at,
                }, content, fname)
                existing_file_labels.setdefault(deal["id"], set()).add(label)
                batch += 1
                if batch % 25 == 0:
                    print(f"  files {batch}")
            except Exception as e:
                report["files_errors"] += 1
                if report["files_errors"] <= 5:
                    print(f"  file error: {e}")


token_global = ""


def main():
    global token_global
    ap = argparse.ArgumentParser()
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--skip-files", action="store_true")
    ap.add_argument("--force-contacts", action="store_true")
    ap.add_argument("--overwrite-fields", action="store_true")
    ap.add_argument("--reimport-tasks", action="store_true", help="Delete Amo-imported tasks before import")
    ap.add_argument("--data-dir", default="/tmp")
    ap.add_argument("--report", default="/tmp/amo-crm-import-report.json")
    args = ap.parse_args()

    data = args.data_dir
    paths = {
        "meta": os.path.join(data, "amo_deals_meta_export.xlsx"),
        "comments": os.path.join(data, "amo_comments_export.xlsx"),
        "open_tasks": os.path.join(data, "amo_open_tasks.xlsx"),
        "stage_history": os.path.join(data, "amo_stage_history.xlsx"),
        "loss_reasons": os.path.join(data, "amo_loss_reasons.xlsx"),
        "attachments_manifest": os.path.join(data, "amo_attachments_manifest.xlsx"),
        "attachments_zip": os.path.join(data, "amo_attachments.zip"),
    }

    for key, p in paths.items():
        if key.endswith("_zip") or args.skip_files and key.startswith("attachments"):
            continue
        if not os.path.isfile(p):
            print(f"MISSING: {p}")
            return 1

    pb_env, email, password = load_env()
    global PB, token_global
    PB = pb_env
    token_global = pb_token(pb_env, email, password)

    print(f"import-amo-crm-data {'APPLY' if args.apply else 'DRY-RUN'} PB={PB}")
    ensure_v4_collections(args.apply)

    deals = pb_list(token_global, "deals")
    amo_map = build_amo_deal_map(deals)
    print(f"deals: {len(deals)}, with amo_id: {len(amo_map)}")

    deal_info_rows = pb_list(token_global, "deal_info")
    deal_info_by_deal = {r["deal"]: r for r in deal_info_rows}

    existing_refs = {r["ref_id"] for r in pb_list(token_global, "deal_activities", 'ref_id!=""', "ref_id") if r.get("ref_id")}

    existing_tasks = pb_list(token_global, "deal_tasks", fields="id,deal,title,due_at,status,description,created_by,done_at")
    if args.reimport_tasks and args.apply:
        deleted = 0
        for t in existing_tasks:
            if (t.get("created_by") or "") == "Amo import":
                req = urllib.request.Request(
                    f"{PB}/api/collections/deal_tasks/records/{t['id']}",
                    headers={"Authorization": token_global}, method="DELETE")
                urllib.request.urlopen(req, timeout=60)
                deleted += 1
        print(f"  reimport-tasks: deleted {deleted} Amo tasks")
        existing_tasks = []
    existing_task_keys = set()
    existing_amo_task_ids = set()
    for t in existing_tasks:
        existing_task_keys.add((t["deal"], t.get("title") or "", (t.get("due_at") or t.get("done_at") or "")[:19]))
        desc = t.get("description") or ""
        m = re.search(r"amo_task_id=(\d+)", desc)
        if m:
            existing_amo_task_ids.add(m.group(1))
        m2 = re.search(r"amo_task_done:(\d+)", desc)
        if m2:
            existing_amo_task_ids.add(f"amo_task_done:{m2.group(1)}")

    existing_contacts = defaultdict(list)
    for c in pb_list(token_global, "deal_contacts", fields="id,deal"):
        existing_contacts[c["deal"]].append(c)

    existing_file_labels = defaultdict(set)
    for f in pb_list(token_global, "deal_files", fields="deal,label"):
        if f.get("label", "").startswith("Amo:"):
            existing_file_labels[f["deal"]].add(f["label"])

    report = defaultdict(int)
    report["questions"] = []

    _, meta_deals_rows = read_xlsx_sheet(paths["meta"], "Сделки")
    meta_deals = {}
    for r in meta_deals_rows:
        try:
            meta_deals[int(r["amo_lead_id"])] = r
        except (TypeError, ValueError, KeyError):
            pass

    _, contacts_rows = read_xlsx_sheet(paths["meta"], "Контакты")
    _, companies_rows = read_xlsx_sheet(paths["meta"], "Компании")
    _, comments_rows = read_xlsx_sheet(paths["comments"], "events")
    _, open_tasks_rows = read_xlsx_sheet(paths["open_tasks"])
    _, stage_rows = read_xlsx_sheet(paths["stage_history"])
    _, manifest_rows = read_xlsx_sheet(paths["attachments_manifest"]) if os.path.isfile(paths["attachments_manifest"]) else ([], [])

    overwrite = args.overwrite_fields

    print("1/6 deal fields + deal_info...")
    import_deal_fields(meta_deals, amo_map, deal_info_by_deal, args.apply, overwrite, report)

    print("2/6 contacts...")
    import_contacts(contacts_rows, amo_map, existing_contacts, args.apply, args.force_contacts, report)

    print("3/6 companies -> deal_info...")
    import_companies(companies_rows, amo_map, deal_info_by_deal, args.apply, overwrite, report)

    print("4/6 activities (comments + stage history)...")
    import_activities(comments_rows, stage_rows, amo_map, existing_refs, args.apply, report)

    print("5/6 tasks (open + completed)...")
    import_tasks(open_tasks_rows, comments_rows, amo_map, existing_task_keys, existing_amo_task_ids, args.apply, report)

    if not args.skip_files:
        print("6/6 attachments...")
        import_files(manifest_rows, paths["attachments_zip"], amo_map, existing_file_labels, args.apply, report)
    else:
        print("6/6 attachments skipped")

    # uncertain fields for user decision (iteration 2)
    col_counts = defaultdict(int)
    for r in meta_deals_rows:
        for k, v in r.items():
            if k in SKIP_DEAL_COLS or v in (None, ""):
                continue
            if not map_field_name(k):
                col_counts[k] += 1
    report["uncertain_fields"] = [{"field": k, "filled_rows": v} for k, v in sorted(col_counts.items(), key=lambda x: -x[1])[:30]]

    report["questions"] = [
        "Куда класть «Продукт ИТМен», «Оборот сделки», «Формат закупки», «Конечные точки» — в deals или отдельные поля?",
        "«Бизнес задачи», «Ссылка на Kaiten», Product_Focus — в pains/capabilities или deal_info.notes?",
        "Теги Amo сейчас пишутся в deal_info.notes — нужна отдельная коллекция/поле?",
        "Нужно ли импортировать loss_reasons.xlsx отдельно (если в meta уже пусто)?",
        "Вложения: label «Amo:{note_id}» — ок или нужна категория (ТЗ/КП/Договор)?",
    ]

    out = {k: (dict(v) if isinstance(v, defaultdict) else v) for k, v in report.items()}
    json.dump(out, open(args.report, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(json.dumps({k: out[k] for k in out if k not in ("uncertain_fields", "questions")}, ensure_ascii=False, indent=2))
    print(f"report: {args.report}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
