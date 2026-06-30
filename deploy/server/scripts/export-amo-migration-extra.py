#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Extra Amo exports: open tasks, stage history, loss reasons, attachments archive."""
from __future__ import annotations

import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import ssl
import zipfile
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook, load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TOKEN_FILE = os.path.join(SCRIPT_DIR, "amo-tokens.json")
SUBDOMAIN = os.environ.get("AMO_SUBDOMAIN", "inferit")
BASE = f"https://{SUBDOMAIN}.amocrm.ru"
CLIENT_ID = os.environ.get("AMO_CLIENT_ID", "f4ae4a8e-f973-406a-906a-fc3e29d4a2d9")
CLIENT_SECRET = os.environ.get("AMO_CLIENT_SECRET", "")
REDIRECT_URI = os.environ.get("AMO_REDIRECT_URI", "https://itmen-pipeline.nwlvl.ru/")


def http_json(url, data=None, token=None, method=None, timeout=180):
    headers = {"Content-Type": "application/json", "User-Agent": "ITMen-amo-export/1.0"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    body = None if data is None else json.dumps(data).encode("utf-8")
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    with urllib.request.urlopen(req, timeout=timeout) as res:
        raw = res.read().decode("utf-8")
        return json.loads(raw) if raw else {}


def http_bytes(url, timeout=300):
    # Drive download URLs are pre-signed (S3-style); Authorization header breaks them (HTTP 400).
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    req = urllib.request.Request(url, headers={"User-Agent": "ITMen-amo-export/1.0"})
    with urllib.request.urlopen(req, timeout=timeout, context=ctx) as res:
        return res.read()


def get_access_token():
    stored = json.load(open(TOKEN_FILE, encoding="utf-8"))
    if stored.get("refresh_token") and CLIENT_SECRET:
        try:
            data = http_json(
                f"{BASE}/oauth2/access_token",
                {
                    "client_id": CLIENT_ID,
                    "client_secret": CLIENT_SECRET,
                    "grant_type": "refresh_token",
                    "refresh_token": stored["refresh_token"],
                    "redirect_uri": REDIRECT_URI,
                },
            )
            data["saved_at"] = datetime.now(timezone.utc).isoformat()
            json.dump(data, open(TOKEN_FILE, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
            return data["access_token"]
        except Exception:
            pass
    return stored["access_token"]


def ts_fmt(val):
    if not val:
        return ""
    if isinstance(val, int):
        return datetime.fromtimestamp(val, timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    return str(val)


def load_lead_ids(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    id_idx = idx.get("ID", 0)
    name_idx = idx.get("Название сделки", 1)
    out = {}
    for row in rows[1:]:
        if not row or row[id_idx] in (None, ""):
            continue
        out[int(row[id_idx])] = str(row[name_idx] or "").strip() if name_idx < len(row) else ""
    return out


def fetch_users(token):
    users = {}
    page = 1
    while True:
        data = http_json(f"{BASE}/api/v4/users?page={page}&limit=250", token=token)
        for u in (data.get("_embedded") or {}).get("users") or []:
            if u.get("id") is not None:
                users[int(u["id"])] = u.get("name") or u.get("login") or str(u["id"])
        if not (data.get("_links") or {}).get("next"):
            break
        page += 1
        time.sleep(0.1)
    return users


def fetch_status_maps(token):
    pipelines = {}
    statuses = {}
    data = http_json(f"{BASE}/api/v4/leads/pipelines", token=token)
    for p in (data.get("_embedded") or {}).get("pipelines") or []:
        if p.get("id") is not None:
            pipelines[int(p["id"])] = p.get("name") or str(p["id"])
        for s in (p.get("_embedded") or {}).get("statuses") or []:
            if s.get("id") is not None:
                statuses[int(s["id"])] = {
                    "name": s.get("name") or str(s["id"]),
                    "pipeline_id": p.get("id"),
                    "pipeline_name": p.get("name") or "",
                }
    return pipelines, statuses


def write_xlsx(path, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    wb.save(path)


def export_open_tasks(token, lead_ids, lead_names, users, out_path):
    now = int(datetime.now(timezone.utc).timestamp())
    rows = []
    lead_list = sorted(lead_ids)
    for i in range(0, len(lead_list), 10):
        batch = lead_list[i:i + 10]
        params = [
            ("filter[entity_type]", "leads"),
            ("filter[is_completed]", "0"),
            ("limit", "250"),
        ]
        for lid in batch:
            params.append(("filter[entity_id][]", str(lid)))
        q = urllib.parse.urlencode(params)
        page = 1
        while True:
            data = http_json(f"{BASE}/api/v4/tasks?{q}&page={page}", token=token)
            for t in (data.get("_embedded") or {}).get("tasks") or []:
                if int(t.get("entity_id") or 0) not in lead_ids:
                    continue
                due = t.get("complete_till") or 0
                overdue = bool(due and due < now)
                rid = t.get("responsible_user_id")
                rows.append({
                    "amo_lead_id": t.get("entity_id"),
                    "Название сделки": lead_names.get(int(t.get("entity_id") or 0), ""),
                    "task_id": t.get("id"),
                    "Задача": (t.get("text") or "").strip(),
                    "Ответственный": users.get(int(rid), rid) if rid else "",
                    "Срок": ts_fmt(due),
                    "Просрочена": "да" if overdue else "",
                    "Создана": ts_fmt(t.get("created_at")),
                    "Обновлена": ts_fmt(t.get("updated_at")),
                })
            if not (data.get("_links") or {}).get("next"):
                break
            page += 1
            time.sleep(0.1)
        time.sleep(0.08)
    rows.sort(key=lambda r: (r["amo_lead_id"], r["Срок"]))
    headers = ["amo_lead_id", "Название сделки", "task_id", "Задача", "Ответственный", "Срок", "Просрочена", "Создана", "Обновлена"]
    write_xlsx(out_path, headers, rows)
    print(f"open tasks: {len(rows)} (overdue {sum(1 for r in rows if r['Просрочена'])}) -> {out_path}")
    return rows


def export_stage_history(token, lead_ids, lead_names, users, statuses, out_path):
    rows = []
    page = 1
    while True:
        data = http_json(
            f"{BASE}/api/v4/events?filter[entity]=lead&filter[type]=lead_status_changed&page={page}&limit=100",
            token=token,
        )
        for ev in (data.get("_embedded") or {}).get("events") or []:
            eid = ev.get("entity_id")
            if eid is None or int(eid) not in lead_ids:
                continue
            before = (ev.get("value_before") or [{}])[0].get("lead_status") or {}
            after = (ev.get("value_after") or [{}])[0].get("lead_status") or {}
            bid, aid = before.get("id"), after.get("id")
            bname = statuses.get(int(bid), {}).get("name", bid) if bid else ""
            aname = statuses.get(int(aid), {}).get("name", aid) if aid else ""
            pipe = statuses.get(int(aid), {}).get("pipeline_name") if aid else ""
            author = users.get(int(ev["created_by"]), ev.get("created_by")) if ev.get("created_by") else ""
            rows.append({
                "amo_lead_id": eid,
                "Название сделки": lead_names.get(int(eid), ""),
                "event_id": ev.get("id"),
                "Дата": ts_fmt(ev.get("created_at")),
                "Автор": author,
                "Воронка": pipe,
                "Было": bname,
                "Стало": aname,
            })
        if not (data.get("_links") or {}).get("next"):
            break
        page += 1
        if page % 20 == 0:
            print(f"  stage events page {page}, matched {len(rows)}")
        time.sleep(0.12)
    rows.sort(key=lambda r: (r["amo_lead_id"], r["Дата"]))
    headers = ["amo_lead_id", "Название сделки", "event_id", "Дата", "Автор", "Воронка", "Было", "Стало"]
    write_xlsx(out_path, headers, rows)
    print(f"stage history: {len(rows)} -> {out_path}")
    return rows


def export_loss_reasons(token, lead_ids, lead_names, users, out_path):
    rows = []
    id_list = sorted(lead_ids)
    for i in range(0, len(id_list), 50):
        batch = id_list[i:i + 50]
        params = [("limit", "250"), ("with", "loss_reason")]
        for lid in batch:
            params.append(("filter[id][]", str(lid)))
        q = urllib.parse.urlencode(params)
        data = http_json(f"{BASE}/api/v4/leads?{q}", token=token)
        for lead in (data.get("_embedded") or {}).get("leads") or []:
            lid = int(lead.get("id") or 0)
            if lid not in lead_ids:
                continue
            loss = (lead.get("_embedded") or {}).get("loss_reason") or []
            if not loss and not lead.get("loss_reason_id"):
                continue
            loss_name = loss[0].get("name") if loss else ""
            rid = lead.get("responsible_user_id")
            rows.append({
                "amo_lead_id": lid,
                "Название сделки": lead.get("name") or lead_names.get(lid, ""),
                "Причина отказа": loss_name,
                "loss_reason_id": lead.get("loss_reason_id") or "",
                "Этап закрытия": lead.get("status_id") or "",
                "Закрыта": ts_fmt(lead.get("closed_at")),
                "Ответственный": users.get(int(rid), rid) if rid else "",
            })
        time.sleep(0.1)
    headers = ["amo_lead_id", "Название сделки", "Причина отказа", "loss_reason_id", "Этап закрытия", "Закрыта", "Ответственный"]
    write_xlsx(out_path, headers, rows)
    print(f"loss reasons: {len(rows)} -> {out_path}")
    return rows


def safe_filename(name, fallback="file"):
    name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", str(name or fallback)).strip()
    return name[:180] or fallback


def export_attachments(token, lead_ids, lead_names, users, manifest_path, zip_path, skip_download=False):
    drive_url = http_json(f"{BASE}/api/v4/account?with=drive_url", token=token).get("drive_url")
    if not drive_url:
        raise RuntimeError("drive_url not available")
    manifest = []
    seen_uuids = set()
    page = 1
    while True:
        data = http_json(f"{BASE}/api/v4/leads/notes?filter[note_type]=attachment&page={page}&limit=250", token=token)
        for note in (data.get("_embedded") or {}).get("notes") or []:
            lid = note.get("entity_id")
            if lid is None or int(lid) not in lead_ids:
                continue
            params = note.get("params") or {}
            fuuid = params.get("file_uuid")
            if not fuuid or fuuid in seen_uuids:
                continue
            seen_uuids.add(fuuid)
            fname = params.get("file_name") or params.get("original_name") or params.get("text") or f"{fuuid}.bin"
            fname = safe_filename(fname, fuuid[:8])
            arc_path = f"{lid}/{note.get('id')}_{fname}"
            author = users.get(int(note.get("created_by") or 0), note.get("created_by") or "")
            manifest.append({
                "amo_lead_id": lid,
                "Название сделки": lead_names.get(int(lid), ""),
                "note_id": note.get("id"),
                "file_uuid": fuuid,
                "version_uuid": params.get("version_uuid") or "",
                "Имя файла": fname,
                "Архивный путь": arc_path,
                "Дата": ts_fmt(note.get("created_at")),
                "Автор": author,
                "download_status": "pending",
            })
        if not (data.get("_links") or {}).get("next"):
            break
        page += 1
        time.sleep(0.12)

    headers = ["amo_lead_id", "Название сделки", "note_id", "file_uuid", "version_uuid", "Имя файла", "Архивный путь", "Дата", "Автор", "download_status"]
    write_xlsx(manifest_path, headers, manifest)
    print(f"attachments manifest: {len(manifest)} -> {manifest_path}")

    if skip_download:
        return manifest

    ok = fail = 0
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for i, row in enumerate(manifest, 1):
            fuuid = row["file_uuid"]
            arc_path = row["Архивный путь"]
            try:
                meta = http_json(f"{drive_url}/v1.0/files/{fuuid}", token=token)
                links = meta.get("_links") or {}
                dl = ((links.get("download_version") or links.get("download")) or {}).get("href")
                if not dl:
                    row["download_status"] = "no_download_link"
                    fail += 1
                    continue
                content = http_bytes(dl)
                zf.writestr(arc_path, content)
                row["download_status"] = "ok"
                row["size_bytes"] = len(content)
                ok += 1
            except Exception as e:
                row["download_status"] = f"error: {e}"[:120]
                fail += 1
            if i % 25 == 0:
                print(f"  attachments {i}/{len(manifest)} ok={ok} fail={fail}")
            time.sleep(0.05)
    write_xlsx(manifest_path, headers + (["size_bytes"] if "size_bytes" not in headers else []), manifest)
    print(f"attachments zip: ok={ok} fail={fail} -> {zip_path}")
    return manifest


def main():
    xlsx_in = sys.argv[1] if len(sys.argv) > 1 else r"c:\Users\Данила\Downloads\amocrm_export_leads_2026-06-25.xlsx"
    out_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.dirname(xlsx_in)
    skip_dl = "--skip-download" in sys.argv

    lead_names = load_lead_ids(xlsx_in)
    lead_ids = set(lead_names.keys())
    print(f"leads: {len(lead_ids)}")

    token = get_access_token()
    users = fetch_users(token)
    _, statuses = fetch_status_maps(token)

    export_open_tasks(token, lead_ids, lead_names, users, os.path.join(out_dir, "amo_open_tasks.xlsx"))
    export_stage_history(token, lead_ids, lead_names, users, statuses, os.path.join(out_dir, "amo_stage_history.xlsx"))
    export_loss_reasons(token, lead_ids, lead_names, users, os.path.join(out_dir, "amo_loss_reasons.xlsx"))
    export_attachments(
        token, lead_ids, lead_names, users,
        os.path.join(out_dir, "amo_attachments_manifest.xlsx"),
        os.path.join(out_dir, "amo_attachments.zip"),
        skip_download=skip_dl,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
