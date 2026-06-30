#!/usr/bin/env python3
"""Parse kaiten-match-export.xlsx → import-plan.json for run-kaiten-import.js"""
import json
import re
import sys
from pathlib import Path

import openpyxl

CARD_RE = re.compile(r"/card[s]?/(\d+)")


def is_red(cell):
    fill = cell.fill
    if not fill:
        return False
    for attr in ("fgColor", "start_color", "end_color"):
        c = getattr(fill, attr, None)
        if c is None:
            continue
        rgb = getattr(c, "rgb", None) or getattr(c, "value", None)
        if rgb and str(rgb).upper().endswith("FF0000"):
            return True
    return False


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Downloads" / "kaiten-match-export.xlsx"
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else xlsx.with_name("kaiten-import-plan.json")

    wb = openpyxl.load_workbook(xlsx)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[i - 1]: ws.cell(r, i).value for i in range(1, ws.max_column + 1)}
        deal_id = row.get("deal_id")
        if not deal_id:
            continue
        url = str(row.get("kaiten_url") or "")
        m = CARD_RE.search(url)
        card_id = int(m.group(1)) if m else None
        if not card_id and row.get("kaiten_card_id"):
            try:
                card_id = int(row.get("kaiten_card_id"))
            except (TypeError, ValueError):
                card_id = None

        red_owner = is_red(ws.cell(r, 4)) if ws.max_column >= 4 else False
        force_owner = str(row.get("presale_owner") or "").strip() if red_owner else ""

        if card_id:
            action = "link"
        else:
            action = "create"

        rows.append({
            "dealId": str(deal_id).strip(),
            "action": action,
            "cardId": card_id,
            "forceOwner": force_owner or None,
            "resetBackfill": True,
        })

    plan = {"rows": rows, "meta": {"source": str(xlsx), "count": len(rows)}}
    out.write_text(json.dumps(plan, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"ok": True, "out": str(out), "linked": sum(1 for x in rows if x["action"] == "link"), "create": sum(1 for x in rows if x["action"] == "create")}, ensure_ascii=False))


if __name__ == "__main__":
    main()
