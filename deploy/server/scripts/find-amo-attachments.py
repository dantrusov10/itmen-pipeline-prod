#!/usr/bin/env python3
import sys
from openpyxl import load_workbook
path = sys.argv[1] if len(sys.argv) > 1 else "/tmp/amo-import/amo_attachments_manifest.xlsx"
needle = sys.argv[2] if len(sys.argv) > 2 else "27524739"
wb = load_workbook(path, read_only=True, data_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
h = [str(c or "").strip() for c in rows[0]]
idx = {x: i for i, x in enumerate(h)}
for r in rows[1:]:
    if not r:
        continue
    lid = str(r[idx["amo_lead_id"]] or "")
    name = str(r[idx.get("Название сделки", 1)] or "")
    if needle in lid or needle.lower() in name.lower() or "втб" in name.lower() and "банк" in name.lower():
        print(lid, name, r[idx.get("Имя файла", 0)], r[idx.get("download_status", 0)], r[idx.get("Архивный путь", 0)])
