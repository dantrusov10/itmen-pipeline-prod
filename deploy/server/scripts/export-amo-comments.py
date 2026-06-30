#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Export AmoCRM lead notes to Excel for leads listed in amo export xlsx."""
from __future__ import annotations

import json
import os
import re
import sys
import time
import urllib.error
import urllib.request
from datetime import datetime

from openpyxl import Workbook, load_workbook

SUBDOMAIN = os.environ.get("AMO_SUBDOMAIN", "inferit")
BASE = f"https://{SUBDOMAIN}.amocrm.ru"
TOKEN_FILE = os.environ.get(
    "AMO_TOKEN_FILE",
    os.path.join(os.path.dirname(__file__), "amo-tokens.json"),
)
CLIENT_ID = os.environ.get("AMO_CLIENT_ID", "f4ae4a8e-f973-406a-906a-fc3e29d4a2d9")
CLIENT_SECRET = os.environ.get("AMO_CLIENT_SECRET", "")
REDIRECT_URI = os.environ.get("AMO_REDIRECT_URI", "https://itmen-pipeline.nwlvl.ru/")
AUTH_CODE = os.environ.get("AMO_AUTH_CODE", "")
ACCESS_TOKEN = os.environ.get("AMO_ACCESS_TOKEN", "")

NOTE_TYPE_LABELS = {
    "common": "Комментарий",
    "call_in": "Входящий звонок",
    "call_out": "Исходящий звонок",
    "sms_in": "Входящее SMS",
    "sms_out": "Исходящее SMS",
    "service_message": "Системное",
    "extended_service_message": "Системное (расш.)",
    "attachment": "Файл",
    "geolocation": "Геолокация",
    "message_cashier": "Сообщение кассиру",
}


def http_json(url, data=None, token=None, method=None):
    headers = {"Content-Type": "application/json", "User-Agent": "ITMen-amo-export/1.0"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    body = None if data is None else json.dumps(data).encode("utf-8")
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=180) as res:
            raw = res.read().decode("utf-8")
            return json.loads(raw) if raw else {}
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {e.code} {url}: {detail}") from e


def save_tokens(data, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump({**data, "saved_at": datetime.utcnow().isoformat() + "Z"}, f, ensure_ascii=False, indent=2)


def load_tokens(path):
    if not path or not os.path.isfile(path):
        return None
    return json.load(open(path, encoding="utf-8"))


def get_access_token():
    global ACCESS_TOKEN
    if ACCESS_TOKEN:
        return ACCESS_TOKEN

    stored = load_tokens(TOKEN_FILE) if TOKEN_FILE else None
    if stored and stored.get("access_token"):
        if stored.get("refresh_token"):
            try:
                refreshed = refresh_access_token(stored["refresh_token"])
                if TOKEN_FILE:
                    save_tokens(refreshed, TOKEN_FILE)
                ACCESS_TOKEN = refreshed["access_token"]
                return ACCESS_TOKEN
            except Exception:
                pass
        ACCESS_TOKEN = stored["access_token"]
        return ACCESS_TOKEN

    if not AUTH_CODE:
        raise RuntimeError("Нужен AMO_AUTH_CODE или AMO_ACCESS_TOKEN")
    data = http_json(
        f"{BASE}/oauth2/access_token",
        {
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "grant_type": "authorization_code",
            "code": AUTH_CODE,
            "redirect_uri": REDIRECT_URI,
        },
    )
    if TOKEN_FILE:
        save_tokens(data, TOKEN_FILE)
    ACCESS_TOKEN = data["access_token"]
    return ACCESS_TOKEN


def refresh_access_token(refresh_token):
    return http_json(
        f"{BASE}/oauth2/access_token",
        {
            "client_id": CLIENT_ID,
            "client_secret": CLIENT_SECRET,
            "grant_type": "refresh_token",
            "refresh_token": refresh_token,
            "redirect_uri": REDIRECT_URI,
        },
    )


def load_leads_from_xlsx(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    rows = list(wb.active.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    id_idx = idx.get("ID", 0)
    name_idx = idx.get("Название сделки", 1)
    company_idx = idx.get("Компания", 2)
    leads = {}
    for row in rows[1:]:
        if not row or row[id_idx] in (None, ""):
            continue
        try:
            lead_id = int(row[id_idx])
        except (TypeError, ValueError):
            continue
        name = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] is not None else ""
        company = str(row[company_idx]).strip() if company_idx < len(row) and row[company_idx] is not None else ""
        leads[lead_id] = {"name": name, "company": company}
    return leads


def paginate(token, path, params=None, embedded_key=None):
    page = 1
    params = dict(params or {})
    while True:
        q = "&".join(f"{k}={urllib.parse.quote(str(v))}" for k, v in params.items())
        url = f"{BASE}{path}?page={page}&limit=250"
        if q:
            url += "&" + q
        data = http_json(url, token=token)
        embedded = data.get("_embedded") or {}
        key = embedded_key
        if not key:
            key = next((k for k in embedded if k.endswith("notes") or k == "users" or k == "tasks"), None)
        if not key:
            for k, v in embedded.items():
                if isinstance(v, list):
                    key = k
                    break
        items = embedded.get(key, []) if key else []
        for item in items:
            yield item
        links = data.get("_links") or {}
        if not links.get("next"):
            break
        page += 1
        time.sleep(0.15)


import urllib.parse  # noqa: E402


def fetch_users(token):
    users = {}
    for u in paginate(token, "/api/v4/users"):
        uid = u.get("id")
        if uid is None:
            continue
        users[int(uid)] = u.get("name") or u.get("login") or str(uid)
    return users


def note_text(note):
    ntype = note.get("note_type") or ""
    params = note.get("params") or {}
    if ntype == "common":
        return (params.get("text") or "").strip()
    if ntype in ("call_in", "call_out"):
        parts = []
        if params.get("phone"):
            parts.append(str(params["phone"]))
        if params.get("duration") is not None:
            parts.append(f"{params['duration']} сек")
        if params.get("link"):
            parts.append(str(params["link"]))
        if params.get("call_result"):
            parts.append(str(params["call_result"]))
        return " | ".join(parts) or ntype
    if ntype in ("sms_in", "sms_out"):
        return (params.get("text") or "").strip() or ntype
    if ntype in ("service_message", "extended_service_message"):
        svc = params.get("service") or ""
        txt = params.get("text") or ""
        return f"{svc}: {txt}".strip(": ")
    if ntype == "attachment":
        fn = params.get("file_name") or params.get("original_name") or ""
        return fn or "Вложение"
    if ntype == "geolocation":
        return (params.get("text") or params.get("address") or "Геолокация").strip()
    return json.dumps(params, ensure_ascii=False) if params else ntype


def ts_fmt(val):
    if not val:
        return ""
    if isinstance(val, int):
        return datetime.utcfromtimestamp(val).strftime("%Y-%m-%d %H:%M:%S")
    return str(val)


def task_result_text(result):
    if isinstance(result, dict):
        return (result.get("text") or "").strip()
    if isinstance(result, list) and result:
        if isinstance(result[0], dict):
            return (result[0].get("text") or "").strip()
        return str(result[0]).strip()
    return ""


def export_notes(token, lead_ids, users, leads_meta):
    rows = []
    total_seen = 0
    matched = 0
    for note in paginate(token, "/api/v4/leads/notes", embedded_key="notes"):
        total_seen += 1
        entity_id = note.get("entity_id")
        if entity_id is None:
            continue
        try:
            lid = int(entity_id)
        except (TypeError, ValueError):
            continue
        if lid not in lead_ids:
            continue
        matched += 1
        created_by = note.get("created_by")
        author = users.get(int(created_by), str(created_by or "—")) if created_by else "—"
        meta = leads_meta.get(lid, {})
        ntype = note.get("note_type") or ""
        rows.append({
            "amo_lead_id": lid,
            "deal_name": meta.get("name") or "",
            "company": meta.get("company") or "",
            "source_id": note.get("id"),
            "event_type": NOTE_TYPE_LABELS.get(ntype, ntype),
            "author": author,
            "created_at": ts_fmt(note.get("created_at")),
            "task_text": "",
            "text": note_text(note),
        })
    print(f"notes scanned: {total_seen}, matched: {matched}, exported rows: {len(rows)}")
    return rows


def export_tasks(token, lead_ids, users, leads_meta):
    rows = []
    total_seen = 0
    matched = 0
    lead_list = sorted(lead_ids)
    batch_size = 10
    for i in range(0, len(lead_list), batch_size):
        batch = lead_list[i:i + batch_size]
        params = [
            ("filter[entity_type]", "leads"),
            ("filter[is_completed]", "1"),
            ("limit", "250"),
        ]
        for lid in batch:
            params.append(("filter[entity_id][]", str(lid)))
        q = urllib.parse.urlencode(params)
        page = 1
        while True:
            url = f"{BASE}/api/v4/tasks?{q}&page={page}"
            data = http_json(url, token=token)
            tasks = (data.get("_embedded") or {}).get("tasks") or []
            for task in tasks:
                total_seen += 1
                if (task.get("entity_type") or "") != "leads":
                    continue
                entity_id = task.get("entity_id")
                if entity_id is None:
                    continue
                try:
                    lid = int(entity_id)
                except (TypeError, ValueError):
                    continue
                if lid not in lead_ids:
                    continue
                result = task_result_text(task.get("result"))
                if not result:
                    continue
                matched += 1
                meta = leads_meta.get(lid, {})
                author_id = task.get("updated_by") or task.get("responsible_user_id") or task.get("created_by")
                author = users.get(int(author_id), str(author_id or "—")) if author_id else "—"
                rows.append({
                    "amo_lead_id": lid,
                    "deal_name": meta.get("name") or "",
                    "company": meta.get("company") or "",
                    "source_id": task.get("id"),
                    "event_type": "Задача (результат)",
                    "author": author,
                    "created_at": ts_fmt(task.get("updated_at") or task.get("created_at")),
                    "task_text": (task.get("text") or "").strip(),
                    "text": result,
                })
            links = data.get("_links") or {}
            if not links.get("next"):
                break
            page += 1
            time.sleep(0.15)
        time.sleep(0.1)
    print(f"tasks scanned: {total_seen}, matched with result: {matched}, exported rows: {len(rows)}")
    return rows


def write_xlsx(rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "events"
    headers = [
        "amo_lead_id", "Название сделки", "Компания", "source_id",
        "Тип", "Автор", "Дата", "Задача", "Текст",
    ]
    ws.append(headers)
    for r in rows:
        ws.append([
            r["amo_lead_id"], r["deal_name"], r["company"], r["source_id"],
            r["event_type"], r["author"], r["created_at"], r["task_text"], r["text"],
        ])
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col[:200]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 80))
        ws.column_dimensions[letter].width = max(10, max_len + 2)
    ws.column_dimensions["H"].width = 50
    ws.column_dimensions["I"].width = 80
    wb.save(out_path)


def main():
    xlsx_in = sys.argv[1] if len(sys.argv) > 1 else r"c:\Users\Данила\Downloads\amocrm_export_leads_2026-06-25.xlsx"
    xlsx_out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        os.path.dirname(xlsx_in), f"amo_comments_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )

    if not CLIENT_SECRET and os.path.isfile(TOKEN_FILE):
        pass  # refresh via saved tokens
    elif not CLIENT_SECRET:
        print("Set AMO_CLIENT_SECRET or ensure amo-tokens.json exists")
        return 1

    leads_meta = load_leads_from_xlsx(xlsx_in)
    print(f"leads in file: {len(leads_meta)}")

    token = get_access_token()
    print("token OK")

    users = fetch_users(token)
    print(f"users: {len(users)}")

    lead_ids = set(leads_meta.keys())
    note_rows = export_notes(token, lead_ids, users, leads_meta)
    task_rows = export_tasks(token, lead_ids, users, leads_meta)
    rows = note_rows + task_rows
    rows.sort(key=lambda r: (r["amo_lead_id"], r["created_at"], r["source_id"] or 0))
    print(f"total rows: {len(rows)} (notes {len(note_rows)}, tasks {len(task_rows)})")
    write_xlsx(rows, xlsx_out)
    print(f"saved: {xlsx_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
