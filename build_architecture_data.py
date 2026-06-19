# -*- coding: utf-8 -*-
"""Полный каталог вендоров из карты архитектуры → architecture-data.js"""
import json
import os
from collections import OrderedDict

BASE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(BASE)
ARCH_XLSX = os.path.join(ROOT, "Enterprise_IT_Architecture_Map_Export_2026-05-13 (4).xlsx")
OUT = os.path.join(BASE, "js", "architecture-data.js")

import openpyxl

def main():
    if not os.path.isfile(ARCH_XLSX):
        print("NOT FOUND:", ARCH_XLSX)
        return
    wb = openpyxl.load_workbook(ARCH_XLSX, data_only=True)
    ws0 = wb[wb.sheetnames[0]]
    zones_map = OrderedDict()
    for r in range(2, ws0.max_row + 1):
        zone = ws0.cell(r, 1).value
        cls = ws0.cell(r, 2).value
        vendor = ws0.cell(r, 3).value
        product = ws0.cell(r, 4).value
        country = ws0.cell(r, 5).value
        oss = ws0.cell(r, 6).value
        if not zone or not cls:
            continue
        zones_map.setdefault(zone, OrderedDict()).setdefault(cls, [])
        entry = {
            "vendor": str(vendor or "").strip(),
            "product": str(product or "").strip(),
            "country": str(country or "").strip(),
            "openSource": str(oss or "").strip(),
        }
        if entry not in zones_map[zone][cls]:
            zones_map[zone][cls].append(entry)

    ws_cls = wb["Справочник классов"]
    class_meta = {}
    for r in range(2, ws_cls.max_row + 1):
        name = ws_cls.cell(r, 2).value
        if name:
            class_meta[name] = {
                "id": ws_cls.cell(r, 1).value,
                "what": ws_cls.cell(r, 4).value or "",
            }

    ws_fn = wb["Функции по классам"]
    functions = {}
    for r in range(2, ws_fn.max_row + 1):
        cls, fn = ws_fn.cell(r, 1).value, ws_fn.cell(r, 2).value
        if cls and fn:
            functions.setdefault(cls, []).append(fn)

    global_map = {}
    taxonomy = []
    for zone, classes in zones_map.items():
        cls_list = []
        for cls_name, vendors in classes.items():
            meta = class_meta.get(cls_name, {})
            seen, uniq_fn = set(), []
            for fn in functions.get(cls_name, []):
                if fn not in seen:
                    seen.add(fn)
                    uniq_fn.append(fn)
            cls_list.append({
                "id": meta.get("id") or cls_name[:40],
                "name": cls_name,
                "what": meta.get("what", ""),
                "functions": uniq_fn[:12],
                "catalog": vendors,
            })
            for v in vendors:
                key = f"{v['vendor']}|||{v['product']}"
                if key not in global_map:
                    global_map[key] = {
                        **v,
                        "key": key,
                        "label": f"{v['vendor']} — {v['product']}",
                        "classes": [],
                    }
                if cls_name not in global_map[key]["classes"]:
                    global_map[key]["classes"].append(cls_name)
        taxonomy.append({"zone": zone, "classes": cls_list})

    catalog = sorted(global_map.values(), key=lambda x: x["label"].lower())
    arch_out = {
        "zones": taxonomy,
        "globalCatalog": catalog,
        "catalogCount": len(catalog),
        "months": ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                   "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"],
        "years": [2026, 2027, 2028],
    }
    with open(OUT, "w", encoding="utf-8") as f:
        f.write("window.ITMEN_ARCHITECTURE = ")
        json.dump(arch_out, f, ensure_ascii=False)
        f.write(";\n")
    print("OK:", len(catalog), "items ->", OUT)

if __name__ == "__main__":
    main()
