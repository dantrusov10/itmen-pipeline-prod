# -*- coding: utf-8 -*-
"""Шаблон Excel для менеджеров — 1 лист = 1 фамилия"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ITMen_Pipeline_Шаблон_менеджеров.xlsx")

MANAGERS = ["Мерлейн", "Ахметшин", "Сироткин", "Кулагин"]

COLS = [
    ("Клиент", 22), ("Отрасль", 18), ("Стадия", 22), ("Ожидаемая сумма", 14),
    ("Ожидаемый бюджет", 14), ("Партнёр", 16), ("Скидка партнёру %", 12),
    ("Скидка клиенту %", 12), ("Плановый период бюджета", 16), ("Статус бюджета", 18),
    ("Месяц согласования", 12), ("Год согласования", 10),
    ("Что ищут", 40), ("As-IS", 45), ("Боли смены", 45), ("Конкуренты", 55),
    ("Ключевые задачи", 40),
    ("% продукта", 10), ("% пилота", 10), ("Срок задачи", 12),
    ("Следующий шаг", 18), ("Комментарий шаг", 25),
    ("Критический риск", 16), ("Комментарий риск", 25),
    ("Общие боли", 30),
    ("Лояльность (0-5)", 14), ("Статус коммита", 14),
]

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)

EXAMPLE = {
    "Клиент": "Пример: Татнефть",
    "Отрасль": "Нефть и газ",
    "Стадия": "Подготовка Пилота",
    "Ожидаемая сумма": 5000000,
    "Ожидаемый бюджет": 6000000,
    "Партнёр": "Нет партнёра",
    "Скидка партнёру %": 0,
    "Скидка клиенту %": 10,
    "Плановый период бюджета": "Q4 2026",
    "Статус бюджета": "Планируется согласование",
    "Месяц согласования": 9,
    "Год согласования": 2026,
    "Что ищут": "cmdb;discovery;itsm",
    "As-IS": "itsm=Naumen|Service Desk;monitoring=Zabbix|Server",
    "Боли смены": "itsm:Нет CMDB, разрозненные процессы;monitoring:Нет связки с ITSM",
    "Конкуренты": "itsm=ServiceNow|ITSM|evaluating||Strong brand;cmdb=BMC|Helix|rejected|Дорого||",
    "Ключевые задачи": "Инвентаризация и CMDB\nПилот POC\nИнтеграция с AD",
    "% продукта": 75,
    "% пилота": 65,
    "Срок задачи": "2026-09-15",
    "Следующий шаг": "pilot_prep",
    "Комментарий шаг": "Согласовать ТЗ пилота до 01.09",
    "Критический риск": "no_budget",
    "Комментарий риск": "Бюджет не утверждён",
    "Общие боли": "Нет единой CMDB",
    "Лояльность (0-5)": 3,
    "Статус коммита": "verbal",
}

INSTR = [
    ["ITMen — шаблон заполнения пайплайна"],
    [""],
    ["Правила:"],
    ["• 1 лист = 1 менеджер (название листа = фамилия)"],
    ["• 1 строка = 1 клиент / сделка"],
    ["• Строка 1 — заголовки, данные с строки 2"],
    ["• «Что ищут»: cmdb;discovery;itsm;service_desk;itam;sam;monitoring"],
    ["• «As-IS»: сегмент=Вендор|Продукт через ;  Пример: itsm=Naumen|SD;cmdb=Excel|—"],
    ["• «Боли смены»: сегмент:текст через ;  Пример: itsm:нет процессов;cmdb:нет SoT"],
    ["• «Конкуренты»: сегмент=Вендор|Продукт|статус|почему_отказ|почему_смотрят|коммент через ;"],
    ["• Статусы конкурентов: reviewed, evaluating, planned, rejected, selected"],
    ["• «Ключевые задачи»: 1 задача = 1 строка в ячейке (Alt+Enter) или через ;"],
    ["• «Лояльность (0-5)»: ТОЛЬКО ручная оценка доверия клиента к нам (0–5)"],
    ["• «Статус коммита»: none, verbal, email, protocol, loi, guarantee, contract"],
    ["• Скоринг «Коммит» в инструменте считается автоматически из «Статус коммита»"],
    ["• «Следующий шаг» и «Критический риск»: id из листа Справочники"],
    ["• После заполнения загрузите файл в инструмент: Импорт Excel"],
]

REF = [
    ["Сегменты (Что ищут)", "cmdb", "discovery", "itsm", "service_desk", "itam", "sam", "monitoring"],
    ["Статус конкурента", "reviewed", "evaluating", "planned", "rejected", "selected"],
    ["Следующий шаг", "intro", "discovery", "budget", "demo", "pilot_prep", "pilot", "proposal", "terms", "contract"],
    ["Риск", "none", "no_budget", "no_lpr", "competitor", "timing", "technical", "other"],
    ["Статус коммита", "none", "verbal", "email", "protocol", "loi", "guarantee", "contract"],
    ["Лояльность", "0", "1", "2", "3", "4", "5", "— только вручную, не из коммита"],
    ["Скоринг коммита (авто)", "none→0", "verbal→2", "email→2", "protocol→3", "loi→4", "guarantee→4", "contract→5"],
]

def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")

def write_manager_sheet(wb, name):
    from openpyxl.utils import get_column_letter
    ws = wb.create_sheet(title=name)
    for i, (title, width) in enumerate(COLS, 1):
        ws.cell(1, i, title)
        ws.column_dimensions[get_column_letter(i)].width = width
    style_header(ws, len(COLS))
    for i, (title, _) in enumerate(COLS, 1):
        val = EXAMPLE.get(title, "")
        cell = ws.cell(2, i, val)
        if title == "Ключевые задачи" and isinstance(val, str):
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.row_dimensions[2].height = 48

def main():
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Инструкция"
    for i, row in enumerate(INSTR, 1):
        ws0.cell(i, 1, row[0] if row else "")
        if i == 1:
            ws0.cell(i, 1).font = Font(bold=True, size=14, color="1E3A5F")
    ws0.column_dimensions["A"].width = 85

    ws_ref = wb.create_sheet("Справочники")
    for r, row in enumerate(REF, 1):
        for c, val in enumerate(row, 1):
            ws_ref.cell(r, c, val)

    for m in MANAGERS:
        write_manager_sheet(wb, m)

    wb.save(OUT)
    print("OK:", OUT)

if __name__ == "__main__":
    main()
