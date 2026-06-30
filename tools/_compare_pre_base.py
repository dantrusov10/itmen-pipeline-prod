import json, re, urllib.request
from copy import deepcopy
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
url = re.search(r'url:\s*"([^"]+)"', (ROOT/"js/gas-config.js").read_text(encoding="utf-8")).group(1)
import sys; sys.path.insert(0, str(ROOT/"tools"))
from reconstruct_truth_state import norm_ts
from restore_max_audit_state import rebuild_max_state, deal_field_count, score_sum

CUTOFF = "2026-06-25 06:40:58"
rows = json.loads(urllib.request.urlopen(url+"?action=auditAll", timeout=300).read())["rows"]
rows_before = [r for r in rows if norm_ts(r[0]) < CUTOFF]
current = json.loads(urllib.request.urlopen(url+"?action=get", timeout=120).read())["state"]
truth = json.loads((ROOT/"tools/truth_state_preview.json").read_text(encoding="utf-8"))

wb = openpyxl.load_workbook(Path(r"c:\Users\Данила\Downloads\инцидент 9-40-58.xlsx"), read_only=True, data_only=True)
incident_ids = {str(r[2]).strip() for r in wb[wb.sheetnames[0]].iter_rows(min_row=2, values_only=True) if r and r[2]}

pre_cur, _, _, _ = rebuild_max_state(rows_before, deepcopy(current))
pre_tru, _, _, _ = rebuild_max_state(rows_before, deepcopy(truth))

pc = {d["id"]: d for d in pre_cur["deals"]}
pt = {d["id"]: d for d in pre_tru["deals"]}
cc = {d["id"]: d for d in current["deals"]}

print("Compare non-incident: current vs pre(truth base) vs pre(current base)")
for did in sorted(incident_ids, key=lambda x: x):
    pass
n = 0
for did in sorted(pc):
    if did in incident_ids:
        continue
    cf_c = deal_field_count(cc.get(did, {}))
    cf_pt = deal_field_count(pt.get(did, {}))
    cf_pc = deal_field_count(pc.get(did, {}))
    sc_c = score_sum(cc.get(did, {}))
    sc_pt = score_sum(pt.get(did, {}))
    if cf_pt != cf_c or sc_pt != sc_c:
        print(f"  {did} cur fc={cf_c} sc={sc_c} | pre(truth) fc={cf_pt} sc={sc_pt} | pre(cur) fc={cf_pc}")
        n += 1
print("diff count", n)
