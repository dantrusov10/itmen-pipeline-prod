#!/usr/bin/env python3
import openpyxl
from collections import Counter
from pathlib import Path
p = Path(__file__).resolve().parent / "audit_max_vs_prod_gaps.xlsx"
wb = openpyxl.load_workbook(p, read_only=True)
ws = wb.active
rows = list(ws.iter_rows(min_row=2, values_only=True))
print("gaps", len(rows))
by_cust = Counter(r[1] for r in rows)
by_field = Counter(r[3] for r in rows)
print("\nTop fields:")
for f, n in by_field.most_common(15):
    print(f"  {n} {f}")
print("\nClients with gaps:")
for c, n in sorted(by_cust.items(), key=lambda x: -x[1])[:25]:
    print(f"  {n} {c[:50]}")
print("\nSample:")
for r in rows[:8]:
    print(r[0], r[1][:30], "|", r[3], "| audit:", str(r[4])[:40])
wb.close()

log = openpyxl.load_workbook(Path(__file__).resolve().parent / "audit_max_completeness_v2.xlsx", read_only=True)["_справка"]
for row in log.iter_rows(min_row=8, values_only=True):
    if row[0]: print(row)
log.parent.close()
