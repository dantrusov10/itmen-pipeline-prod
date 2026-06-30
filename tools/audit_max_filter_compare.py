#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
1. Сводка «макс. полнота из аудита» с фильтрацией дублей
2. Сверка столбцов L+ (Вероятность …) с продом — подсветка пробелов

Фильтры:
- Softline (клиент) — удалить все строки
- Дубли — оставить одну наиболее заполненную:
  Норникель, ГПН (Газпром Нефть), Петровский завод, прочие по norm(customer)
"""
import json
import re
import sys
import unicodedata
import urllib.request
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "tools"))

from audit_max_export import (  # noqa: E402
    PASSPORT_LABELS,
    collect_best,
    fetch_audit,
    fetch_server_deals,
    field_richness,
    format_cell_value,
    norm_customer,
)
from reconstruct_truth_state import fmt_val, is_empty_audit  # noqa: E402

URL = re.search(
    r'url:\s*"([^"]+)"',
    (ROOT / "js" / "gas-config.js").read_text(encoding="utf-8"),
).group(1)
AMOCRM = Path(r"c:\Users\Данила\Downloads\amocrm_export_leads_2026-06-19 (3).xlsx")
OUT = ROOT / "tools" / "audit_max_completeness_v2.xlsx"
OUT_COMPARE = ROOT / "tools" / "audit_max_vs_prod_gaps.xlsx"

# D–K = первые 8 полей паспорта (импорт AmoCRM, не сверяем)
SKIP_COMPARE = PASSPORT_LABELS[:8]
COMPARE_LABELS = PASSPORT_LABELS[8:]  # L → AD

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
META_FILL = PatternFill("solid", fgColor="E8EEF4")
GAP_FILL = PatternFill("solid", fgColor="FFC7CE")
GAP_FONT = Font(color="9C0006")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")

# Явные группы дублей (нормализованный ключ → подпись)
GROUP_RULES = [
    (re.compile(r"норильск|норникель"), "grp_nornikel"),
    (re.compile(r"газпром\s*нефть"), "grp_gpn"),
    (re.compile(r"петровск"), "grp_petrovsk"),
    (re.compile(r"вкусно"), "grp_vkusno"),
]


def duplicate_group_key(customer: str) -> str:
    c = norm_customer(customer)
    for rx, key in GROUP_RULES:
        if rx.search(c):
            return key
    return f"cust:{c}"


def is_softline_customer(customer: str) -> bool:
    c = norm_customer(customer)
    return c == "softline" or c.startswith("softline ") or " softline" in c


def row_fill_score(row: dict) -> int:
    return sum(
        1 for lbl in PASSPORT_LABELS
        if row["cells"].get(lbl, {}).get("text", "—") not in ("", "—")
    )


def filter_deal_rows(table_rows: list) -> tuple[list, list]:
    """Вернуть (оставшиеся, журнал исключений)."""
    log = []
    # 1) убрать Softline
    kept = []
    for row in table_rows:
        if is_softline_customer(row["customer"]):
            log.append((row["deal_id"], row["customer"], "удалён: Softline"))
            continue
        kept.append(row)

    # 2) дедуп по группе — оставить max fill
    by_group = defaultdict(list)
    for row in kept:
        gk = duplicate_group_key(row["customer"])
        by_group[gk].append(row)

    final = []
    for gk, rows in by_group.items():
        if len(rows) == 1:
            final.append(rows[0])
            continue
        best = max(rows, key=lambda r: (row_fill_score(r), r["deal_id"]))
        for r in rows:
            if r["deal_id"] != best["deal_id"]:
                log.append((
                    r["deal_id"], r["customer"],
                    f"дубль группы {gk}: оставлен {best['deal_id']}",
                ))
        final.append(best)

    final.sort(key=lambda r: int(r["deal_id"].split("-")[1]))
    return final, log


def deal_field_formatted(deal: dict, label: str) -> str:
    raw = fmt_val(deal, label)
    if label == "Скоринг":
        if is_empty_audit(label, raw):
            return ""
        return format_cell_value(label, raw)
    if label == "Риски":
        if not str(raw or "").strip():
            return ""
        return str(raw)
    if is_empty_audit(label, raw):
        return ""
    return format_cell_value(label, raw)


def norm_text(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def richness_from_text(label: str, text: str) -> float:
    if not text or text == "—":
        return 0.0
    return field_richness(label, text)


def norm_task_due(s: str) -> str:
    m = re.search(r"(\d{4}-\d{2}-\d{2})", str(s or ""))
    return m.group(1) if m else ""


def is_gap(audit_text: str, prod_text: str, label: str) -> bool:
    if not audit_text or audit_text == "—":
        return False
    if label == "Срок задачи":
        a, p = norm_task_due(audit_text), norm_task_due(prod_text)
        if not a:
            return False
        # в проде нет даты, в аудите была — пробел; расхождение на 1 день (TZ) не считаем
        return not p
    if not prod_text or prod_text == "—":
        return True
    ar, pr = richness_from_text(label, audit_text), richness_from_text(label, prod_text)
    if ar <= 0:
        return False
    if pr < ar * 0.9:
        return True
    if ar > 3 and norm_text(audit_text) != norm_text(prod_text) and pr < ar:
        return True
    return False


def build_table():
    rows = fetch_audit()
    deals, best = collect_best(rows)
    for d in fetch_server_deals():
        did = str(d.get("id") or "").strip()
        if not did:
            continue
        if did not in deals:
            deals[did] = {"customer": d.get("customer") or "", "owner": d.get("owner") or ""}
        else:
            deals[did]["customer"] = deals[did].get("customer") or d.get("customer") or ""
            deals[did]["owner"] = deals[did].get("owner") or d.get("owner") or ""

    table = []
    for deal_id in sorted(deals.keys(), key=lambda x: int(x.split("-")[1])):
        meta = deals[deal_id]
        row = {
            "deal_id": deal_id,
            "manager": meta.get("owner") or "",
            "customer": meta.get("customer") or "",
            "cells": {},
        }
        for label in PASSPORT_LABELS:
            key = (deal_id, label)
            if key in best:
                _, ts, raw, source = best[key]
                row["cells"][label] = {
                    "text": format_cell_value(label, raw),
                    "comment": f"{ts}\n{source}",
                }
            else:
                row["cells"][label] = {"text": "—", "comment": ""}
        table.append(row)
    return table, len(rows)


def write_main_sheet(wb, table_rows, gaps_map):
    ws = wb.active
    ws.title = "Макс. полнота из аудита"
    fixed = ["ID сделки", "Менеджер", "Клиент"]
    headers = fixed + PASSPORT_LABELS
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(1, col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_index = {label: 4 + i for i, label in enumerate(PASSPORT_LABELS)}

    for r_idx, row in enumerate(table_rows, start=2):
        ws.cell(r_idx, 1, row["deal_id"])
        ws.cell(r_idx, 2, row["manager"])
        ws.cell(r_idx, 3, row["customer"])
        for col in range(1, 4):
            ws.cell(r_idx, col).fill = META_FILL
        for label in PASSPORT_LABELS:
            c_idx = col_index[label]
            text = row["cells"][label]["text"]
            cell = ws.cell(r_idx, c_idx, text)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row["cells"][label]["comment"]:
                cell.comment = Comment(row["cells"][label]["comment"], "аудит")
            if (row["deal_id"], label) in gaps_map:
                cell.fill = GAP_FILL
                cell.font = GAP_FONT
                g = gaps_map[(row["deal_id"], label)]
                extra = f"\n\nПРОД: {g['prod'][:200]}" if g.get("prod") else "\n\nПРОД: пусто"
                if cell.comment:
                    cell.comment = Comment(cell.comment.text + extra, "аудит")
                else:
                    cell.comment = Comment(f"Нет в проде{extra}", "сверка")

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(table_rows) + 1}"
    widths = {1: 10, 2: 22, 3: 32}
    for i, label in enumerate(PASSPORT_LABELS, start=4):
        widths[i] = 28 if label in ("Скоринг", "Ключевые боли", "Конкуренты") else 18
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def write_gaps_sheet(wb, gaps: list):
    ws = wb.create_sheet("Пробелы vs прод")
    ws.append([
        "ID", "Клиент", "Менеджер", "Поле",
        "В аудите (макс.)", "В проде сейчас", "Источник аудита",
    ])
    for col in range(1, 7):
        ws.cell(1, col).fill = HEADER_FILL
        ws.cell(1, col).font = HEADER_FONT
    for g in gaps:
        ws.append([
            g["deal_id"], g["customer"], g["manager"], g["label"],
            g["audit"], g["prod"] or "—", g.get("comment", ""),
        ])
        r = ws.max_row
        for c in range(1, 7):
            ws.cell(r, c).fill = GAP_FILL
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 45
    ws.column_dimensions["G"].width = 22


def write_log_sheet(wb, log, audit_count, n_before, n_after, n_gaps):
    ws = wb.create_sheet("_справка")
    ws.append(["Параметр", "Значение"])
    ws.append(["Строк аудита", audit_count])
    ws.append(["Сделок до фильтра", n_before])
    ws.append(["Сделок после фильтра", n_after])
    ws.append(["Пробелов L+ vs прод", n_gaps])
    ws.append(["Сверка", "столбцы L–AD (Вероятность … Задачи проекта); D–K не сверяются"])
    ws.append([])
    ws.append(["Исключённые / дедуп", ""])
    for item in log:
        ws.append(list(item))


def main():
    print("Сборка таблицы из аудита…")
    table, audit_count = build_table()
    n_before = len(table)
    table, log = filter_deal_rows(table)
    n_after = len(table)
    print(f"Фильтр: {n_before} -> {n_after} сделок, исключено/дедуп: {len(log)}")

    prod_deals = {d["id"]: d for d in fetch_server_deals()}
    gaps = []
    gaps_map = {}

    for row in table:
        did = row["deal_id"]
        deal = prod_deals.get(did)
        if not deal:
            continue
        for label in COMPARE_LABELS:
            audit_text = row["cells"][label]["text"]
            prod_text = deal_field_formatted(deal, label) or "—"
            if is_gap(audit_text, prod_text, label):
                entry = {
                    "deal_id": did,
                    "customer": row["customer"],
                    "manager": row["manager"],
                    "label": label,
                    "audit": audit_text,
                    "prod": prod_text if prod_text != "—" else "",
                    "comment": row["cells"][label]["comment"],
                }
                gaps.append(entry)
                gaps_map[(did, label)] = entry

    print(f"Пробелов в проде (L+): {len(gaps)} по {len({g['deal_id'] for g in gaps})} сделкам")

    wb = Workbook()
    write_main_sheet(wb, table, gaps_map)
    write_gaps_sheet(wb, gaps)
    write_log_sheet(wb, log, audit_count, n_before, n_after, len(gaps))
    wb.save(OUT)
    print(f"Готово: {OUT}")

    # отдельный файл только пробелы
    wb2 = Workbook()
    wb2.remove(wb2.active)
    write_gaps_sheet(wb2, gaps)
    wb2.save(OUT_COMPARE)
    print(f"Пробелы: {OUT_COMPARE}")


if __name__ == "__main__":
    main()
