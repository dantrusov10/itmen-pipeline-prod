#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Восстановление 25.06.2026 09:40:58 МСК:
- 66 сделок из Excel → только поля «Было» из инцидента
- 152 остальных → полная сделка как в аудите ДО инцидента (хронологический replay)
"""
import json
import re
import sys
import urllib.request
from copy import deepcopy
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
URL = re.search(
    r'url:\s*"([^"]+)"',
    (ROOT / "js" / "gas-config.js").read_text(encoding="utf-8"),
).group(1)
EXCEL = Path(r"c:\Users\Данила\Downloads\инцидент 9-40-58.xlsx")
CUTOFF = "2026-06-25 06:40:58"

sys.path.insert(0, str(ROOT / "tools"))
from reconstruct_truth_state import norm_ts  # noqa: E402
from restore_max_audit_state import deal_field_count, score_sum  # noqa: E402
from rollback_burst import apply_field, fetch, norm, post  # noqa: E402


def load_excel_plan():
    wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    plan = []
    incident_ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[2]:
            continue
        deal_id = str(row[2]).strip()
        incident_ids.add(deal_id)
        plan.append({
            "dealId": deal_id,
            "label": str(row[6] or "").strip(),
            "bilo": row[7],
        })
    return incident_ids, plan


def replay_before_incident(rows_before, base_state):
    """Состояние на момент непосредственно до инцидента."""
    state = deepcopy(base_state)
    deal_map = {d["id"]: deepcopy(d) for d in state.get("deals", []) if d.get("id")}
    for row in sorted(rows_before, key=lambda r: (norm_ts(r[0]), str(r[2] or ""), str(r[6] or ""))):
        deal_id = str(row[2] or "")
        label = str(row[6] or "")
        new_val = row[8]
        if not deal_id or not label or label == "—":
            continue
        if deal_id not in deal_map:
            deal_map[deal_id] = {
                "id": deal_id,
                "customer": str(row[3] or ""),
                "owner": str(row[5] or ""),
                "techResearch": {},
                "scores": {},
            }
        apply_field(deal_map[deal_id], label, new_val)
    state["deals"] = sorted(deal_map.values(), key=lambda d: d.get("id", ""))
    return state, deal_map


def main():
    apply = "--apply" in sys.argv
    incident_ids, excel_plan = load_excel_plan()
    print(f"Incident: {len(incident_ids)} deals, {len(excel_plan)} field rows")

    rows = fetch("?action=auditAll").get("rows") or []
    rows_before = [r for r in rows if norm_ts(r[0]) < CUTOFF]
    print(f"Audit rows before {CUTOFF}: {len(rows_before)} / {len(rows)}")

    current = fetch("?action=get")["state"]
    pre_state, pre_map = replay_before_incident(rows_before, current)
    cur_map = {d["id"]: d for d in current.get("deals", []) if d.get("id")}

    final_map = {}
    # 152 нетронутых — целиком из pre-incident replay
    for did, deal in pre_map.items():
        if did not in incident_ids:
            final_map[did] = deepcopy(deal)

    # 66 из инцидента — скелет pre-incident + точечно «Было» из Excel
    for did in incident_ids:
        final_map[did] = deepcopy(pre_map.get(did) or cur_map.get(did) or {"id": did})

    excel_n = 0
    for p in excel_plan:
        deal = final_map.get(p["dealId"])
        if deal and apply_field(deal, p["label"], p["bilo"]):
            excel_n += 1

    final_state = deepcopy(current)
    final_state["deals"] = sorted(final_map.values(), key=lambda d: d.get("id", ""))

    ch_inc = ch_non = 0
    for did, fd in final_map.items():
        cd = cur_map.get(did)
        if not cd:
            continue
        if norm(json.dumps(fd, default=str)) == norm(json.dumps(cd, default=str)):
            continue
        if did in incident_ids:
            ch_inc += 1
        else:
            ch_non += 1

    print(f"Excel fields applied: {excel_n}")
    print(f"Deals to update: incident={ch_inc}, non-incident={ch_non}")

    if ch_non:
        print("Non-incident changes:")
        for did in sorted(final_map):
            if did in incident_ids:
                continue
            cd, fd = cur_map.get(did), final_map.get(did)
            if cd and fd and norm(json.dumps(fd, default=str)) != norm(json.dumps(cd, default=str)):
                print(f"  {did} fc {deal_field_count(cd)}->{deal_field_count(fd)} sc {score_sum(cd)}->{score_sum(fd)}")

    if not apply:
        print("\nPreview. Run with --apply")
        return

  # maintenance on
    post({"action": "setMaintenance", "on": True, "allowMaintenance": True})
    res = post({
        "action": "save",
        "state": final_state,
        "forceFull": True,
        "savedBy": "recover-merge-0940-v2",
        "allowMaintenance": True,
    })
    post({"action": "setMaintenance", "on": False, "allowMaintenance": True})
    if res.get("error"):
        print("ERROR:", res["error"])
        sys.exit(1)
    print("OK updatedAt=", res.get("updatedAt"), "auditRows=", res.get("auditRows"))


if __name__ == "__main__":
    main()
