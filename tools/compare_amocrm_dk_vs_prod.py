#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Сверка D–K (импорт AmoCRM / initial-data.js) с текущим продом."""
import json
import re
import unicodedata
import urllib.request
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
URL = re.search(
    r'url:\s*"([^"]+)"',
    (ROOT / "js" / "gas-config.js").read_text(encoding="utf-8"),
).group(1)
OUT = ROOT / "tools" / "amocrm_dk_vs_prod_gaps.xlsx"

DK_FIELDS = [
    ("Клиент", "customer", "str"),
    ("Отрасль", "industry", "str"),
    ("Стадия", "stage", "str"),
    ("Ожид. сумма", "amount", "num"),
    ("Ожид. бюджет", "expectedBudget", "num"),
    ("Партнёр", "partner", "str"),
    ("Скидка партнёру, %", "partnerDiscount", "num"),
    ("Скидка клиенту, %", "clientDiscount", "num"),
]

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
GAP_FILL = PatternFill("solid", fgColor="FFC7CE")
GAP_FONT = Font(color="9C0006")
OK_FILL = PatternFill("solid", fgColor="C6EFCE")


def norm_str(s):
    s = unicodedata.normalize("NFKC", str(s or "")).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def fmt_num(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def fmt_val(deal, key, kind):
    v = deal.get(key)
    if kind == "num":
        n = fmt_num(v)
        return int(n) if n == int(n) else n
    return str(v or "").strip()


def values_match(kind, a, b, key):
    if kind == "num":
        return abs(fmt_num(a) - fmt_num(b)) < 0.01
    sa, sb = norm_str(a), norm_str(b)
    if sa == sb:
        return True
    # партнёр: нет / пусто
    if key == "partner":
        empty = {"", "нет", "нет партнёра", "—", "-"}
        if sa in empty and sb in empty:
            return True
    return False


def load_initial():
    text = (ROOT / "js" / "initial-data.js").read_text(encoding="utf-8")
    m = re.search(r"window\.ITMEN_INITIAL\s*=\s*(\{.*\})\s*;", text, re.S)
    return json.loads(m.group(1))


def fetch_prod():
    with urllib.request.urlopen(URL + "?action=get", timeout=120) as r:
        return json.loads(r.read().decode("utf-8"))["state"]["deals"]


def main():
    init = load_initial()
    init_map = {d["id"]: d for d in init["deals"]}
    prod_map = {d["id"]: d for d in fetch_prod()}

    gaps = []
    match_count = 0
    total_checks = 0

    for did in sorted(init_map.keys(), key=lambda x: int(x.split("-")[1])):
        ini = init_map[did]
        prod = prod_map.get(did)
        if not prod:
            gaps.append({
                "deal_id": did, "customer": ini.get("customer", ""),
                "owner": ini.get("owner", ""), "field": "—",
                "label": "Сделка", "initial": did, "prod": "НЕТ НА ПРОДЕ",
            })
            continue
        for label, key, kind in DK_FIELDS:
            total_checks += 1
            iv = fmt_val(ini, key, kind)
            pv = fmt_val(prod, key, kind)
            if values_match(kind, iv, pv, key):
                match_count += 1
            else:
                gaps.append({
                    "deal_id": did,
                    "customer": prod.get("customer") or ini.get("customer", ""),
                    "owner": prod.get("owner") or "",
                    "field": key,
                    "label": label,
                    "initial": iv,
                    "prod": pv,
                })

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Пробелы D-K"
    headers = ["ID", "Клиент", "Менеджер", "Поле", "AmoCRM (импорт)", "Прод сейчас"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for g in gaps:
        ws.append([
            g["deal_id"], g["customer"], g["owner"],
            g["label"], g["initial"], g["prod"],
        ])
        r = ws.max_row
        for c in range(1, 7):
            ws.cell(r, c).fill = GAP_FILL
            ws.cell(r, c).font = GAP_FONT

    for col, w in zip("ABCDEF", [10, 36, 22, 22, 28, 28]):
        ws.column_dimensions[col].width = w

    # сводка
    info = wb.create_sheet("_справка")
    info.append(["Параметр", "Значение"])
    info.append(["Сделок в импорте", len(init_map)])
    info.append(["Сделок на проде", len(prod_map)])
    info.append(["Проверок полей D-K", total_checks])
    info.append(["Совпадений", match_count])
    info.append(["Расхождений", len(gaps)])
    info.append(["Поля", "Клиент, Отрасль, Стадия, Сумма, Бюджет, Партнёр, 2 скидки"])
    info.append(["Источник импорта", "js/initial-data.js (AmoCRM 2026-06-19)"])

    # по полям
    from collections import Counter
    by_field = Counter(g["label"] for g in gaps if g["label"] != "Сделка")
    info.append([])
    info.append(["Расхождения по полям", ""])
    for f, n in by_field.most_common():
        info.append([f, n])

    wb.save(OUT)

    print(f"Сделок: {len(init_map)}, проверок: {total_checks}")
    print(f"Совпадений: {match_count}, расхождений: {len(gaps)}")
    print(f"Файл: {OUT}")
    if gaps:
        print("\nПримеры (до 15):")
        for g in gaps[:15]:
            print(f"  {g['deal_id']} {g['customer'][:30]} | {g['label']}: {g['initial']!r} -> {g['prod']!r}")


if __name__ == "__main__":
    main()
