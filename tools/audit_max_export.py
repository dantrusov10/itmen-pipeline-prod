#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Сводная таблица «максимальная полнота из аудита» → Excel.
По каждому полю паспорта — самое полное значение за всю историю (БЫЛО + СТАЛО).
Комментарий к ячейке: дата/время аудита и источник (БЫЛО/СТАЛО).
"""
import json
import re
import unicodedata
import urllib.request
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
URL = re.search(
    r'url:\s*"([^"]+)"',
    (ROOT / "js" / "gas-config.js").read_text(encoding="utf-8"),
).group(1)
OUT = ROOT / "tools" / "audit_max_completeness.xlsx"

# Порядок полей паспорта (метки как в листе _audit)
PASSPORT_LABELS = [
    "Клиент", "Отрасль", "Стадия",
    "Ожид. сумма", "Ожид. бюджет", "Партнёр",
    "Скидка партнёру, %", "Скидка клиенту, %",
    "Вероятность", "Срок задачи",
    "Срок бюджета", "Статус бюджета",
    "Месяц согласования", "Год согласования", "Статус коммита",
    "Ключевые боли", "Риски", "Комментарий к риску",
    "Скоринг",
    "Что ищут", "Другое (что ищут)",
    "% требований проекта", "% требований пилота",
    "Что есть сейчас", "Почему меняют", "Конкуренты", "Задачи проекта",
]

SCORE_NAMES = {
    "loyalty": "Лояльность",
    "commit": "Коммит",
    "budget": "Определённость бюджета",
    "fit": "Соответствие",
    "timing": "Срочность",
    "competitive": "Конкурентная позиция",
    "access": "Доступ к ЛПР",
    "technical": "Техническая готовность",
    "commercial": "Коммерческая готовность",
}

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
META_FILL = PatternFill("solid", fgColor="E8EEF4")


def fetch_audit():
    with urllib.request.urlopen(URL + "?action=auditAll", timeout=300) as r:
        return json.loads(r.read().decode("utf-8")).get("rows") or []


def fetch_server_deals():
    with urllib.request.urlopen(URL + "?action=get", timeout=300) as r:
        state = json.loads(r.read().decode("utf-8")).get("state") or {}
    return state.get("deals") or []


def norm_ts(raw):
    return str(raw or "").strip().replace("T", " ")[:19]


def norm_customer(s):
    s = unicodedata.normalize("NFKC", str(s or "")).lower().strip()
    s = re.sub(r"\s+", " ", s)
    return s


def is_empty_audit(label, raw):
    if raw is None:
        return True
    s = str(raw).strip()
    if not s:
        return True
    if label == "Статус бюджета" and s == "Неизвестно":
        return True
    if label == "Статус коммита" and s in ("none", "Нет подтверждения"):
        return True
    if label == "Срок бюджета" and s == "Не определён":
        return True
    if label == "Ключевые боли" and len(s) < 5:
        return True
    if label == "Скоринг":
        try:
            return sum(json.loads(s).values()) <= 2
        except Exception:
            return True
    if label in ("Риски", "Что ищут") and len(s) < 2:
        return True
    if label in ("Что есть сейчас", "Почему меняют", "Конкуренты") and s in ("{}", ""):
        return True
    return False


def score_sum_raw(raw):
    try:
        return sum(json.loads(str(raw or "{}")).values())
    except Exception:
        return 0


def field_richness(label, raw):
    if is_empty_audit(label, raw):
        return 0.0
    s = str(raw).strip()
    if label == "Скоринг":
        return float(score_sum_raw(raw))
    if label in ("Что есть сейчас", "Почему меняют", "Конкуренты"):
        try:
            obj = json.loads(s)
            if isinstance(obj, dict):
                return float(sum(len(str(v)) for v in obj.values()) + len(obj) * 5)
        except Exception:
            pass
    if label in ("Риски", "Что ищут", "Ключевые боли"):
        return float(len(s))
    if label == "Задачи проекта":
        return float(len([x for x in s.split(";") if x.strip()]))
    if label in ("% требований проекта", "% требований пилота", "Вероятность"):
        try:
            return float(s)
        except Exception:
            return 1.0
    return float(len(s))


def format_scores(raw):
    try:
        d = json.loads(str(raw))
    except Exception:
        return str(raw or "")
    parts = []
    for k, name in SCORE_NAMES.items():
        v = d.get(k, 0) or 0
        if v:
            parts.append(f"{name}: {v}")
    total = sum(d.values())
    tail = f" (Σ={int(total)})" if total else ""
    return "; ".join(parts) + tail if parts else str(raw)


def format_asis(raw):
    try:
        stack = json.loads(str(raw))
    except Exception:
        return str(raw or "")
    if not isinstance(stack, dict):
        return str(raw)
    lines = []
    for seg, entry in stack.items():
        if not entry:
            continue
        if isinstance(entry, str):
            lines.append(f"{seg}: {entry}")
        elif isinstance(entry, dict):
            v = entry.get("vendor") or ""
            p = entry.get("product") or ""
            c = entry.get("comment") or ""
            chunk = " | ".join(x for x in [v, p, c] if x)
            if chunk:
                lines.append(f"{seg}: {chunk}")
    return "\n".join(lines)


def format_change_pains(raw):
    try:
        pains = json.loads(str(raw))
    except Exception:
        return str(raw or "")
    if not isinstance(pains, dict):
        return str(raw)
    return "\n".join(f"{k}: {v}" for k, v in pains.items() if str(v or "").strip())


def format_competitors(raw):
    try:
        entries = json.loads(str(raw))
    except Exception:
        return str(raw or "")
    if not isinstance(entries, dict):
        return str(raw)
    lines = []
    for seg, arr in entries.items():
        for e in arr or []:
            if not isinstance(e, dict):
                continue
            if not (e.get("vendor") or e.get("product")):
                continue
            lines.append(
                f"{seg}: {e.get('vendor', '')} / {e.get('product', '')}"
                + (f" [{e.get('status', '')}]" if e.get("status") else "")
            )
    return "\n".join(lines)


def format_cell_value(label, raw):
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    if label == "Скоринг":
        return format_scores(raw)
    if label == "Что есть сейчас":
        return format_asis(raw)
    if label == "Почему меняют":
        return format_change_pains(raw)
    if label == "Конкуренты":
        return format_competitors(raw)
    if label == "Задачи проекта":
        return "\n".join(x.strip() for x in s.split(";") if x.strip())
    if label == "Что ищут":
        return s.replace(",", "; ")
    return s


def collect_best(rows):
    """deal_id -> meta + label -> (rich, ts, raw, source)."""
    deals = {}
    best = {}

    for row in rows:
        ts = norm_ts(row[0])
        deal_id = str(row[2] or "").strip()
        if not deal_id:
            continue
        customer = str(row[3] or "").strip()
        owner = str(row[4] or "").strip()  # колонка 5 в Excel — ответственный
        if deal_id not in deals:
            deals[deal_id] = {"customer": customer, "owner": owner}
        if customer:
            deals[deal_id]["customer"] = customer
        if owner:
            deals[deal_id]["owner"] = owner

        label = str(row[6] or "")
        if not label or label == "—":
            continue

        for source, raw in (("БЫЛО", row[7]), ("СТАЛО", row[8])):
            r = field_richness(label, raw)
            if r <= 0:
                continue
            key = (deal_id, label)
            cur = best.get(key)
            if cur is None or r > cur[0] or (r == cur[0] and ts >= cur[1]):
                best[key] = (r, ts, raw, source)

    return deals, best


def find_duplicate_customers(deals):
    by_cust = defaultdict(list)
    for did, meta in deals.items():
        c = norm_customer(meta.get("customer"))
        if c:
            by_cust[c].append((did, meta.get("customer")))
    dups = {c: ids for c, ids in by_cust.items() if len(ids) > 1}
    return dups


def build_rows(deals, best):
    out = []
    for deal_id in sorted(deals.keys(), key=lambda x: int(x.split("-")[1]) if "-" in x else x):
        meta = deals[deal_id]
        row = {
            "deal_id": deal_id,
            "manager": meta.get("owner") or "",
            "customer": meta.get("customer") or "",
            "cells": {},
        }
        for label in PASSPORT_LABELS:
            key = (deal_id, label)
            if key in best:
                _, ts, raw, source = best[key]
                row["cells"][label] = {
                    "text": format_cell_value(label, raw),
                    "comment": f"{ts}\n{source}",
                }
            else:
                row["cells"][label] = {"text": "—", "comment": ""}
        out.append(row)
    return out


def write_excel(table_rows, dup_report, audit_count):
    wb = Workbook()
    ws = wb.active
    ws.title = "Макс. полнота из аудита"

    fixed = ["ID сделки", "Менеджер", "Клиент"]
    headers = fixed + PASSPORT_LABELS
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(1, col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {1: 10, 2: 22, 3: 32}
    for i, label in enumerate(PASSPORT_LABELS, start=4):
        widths[i] = 28 if label in ("Скоринг", "Ключевые боли", "Конкуренты") else 18

    for r_idx, row in enumerate(table_rows, start=2):
        ws.cell(r_idx, 1, row["deal_id"])
        ws.cell(r_idx, 2, row["manager"])
        ws.cell(r_idx, 3, row["customer"])
        for c_idx, label in enumerate(PASSPORT_LABELS, start=4):
            cell = ws.cell(r_idx, c_idx, row["cells"][label]["text"])
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            comment_text = row["cells"][label]["comment"]
            if comment_text:
                cell.comment = Comment(comment_text, "аудит")
        for col in range(1, 4):
            ws.cell(r_idx, col).fill = META_FILL

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(table_rows) + 1}"
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Лист с дублями и справкой
    info = wb.create_sheet("_справка")
    info.append(["Параметр", "Значение"])
    info.append(["Строк аудита", audit_count])
    info.append(["Сделок в таблице", len(table_rows)])
    info.append(["Логика", "По каждому полю — самое полное значение из всего аудита (БЫЛО и СТАЛО)"])
    info.append(["Подсказка", "Наведите на ячейку — дата/время и БЫЛО или СТАЛО"])
    info.append(["Менеджер", "Колонка 5 листа _audit (ответственный / owner)"])
    info.append([])
    info.append(["Дубли клиентов (нормализованное имя → deal_id)", ""])
    if dup_report:
        for cust_norm, items in sorted(dup_report.items()):
            info.append([cust_norm, ", ".join(f"{did} ({name})" for did, name in items)])
    else:
        info.append(["—", "дублей не найдено"])

    wb.save(OUT)


def main():
    print("Загрузка аудита…")
    rows = fetch_audit()
    print(f"Строк аудита: {len(rows)}")
    deals, best = collect_best(rows)
    # все сделки пайплайна (даже без строк в аудите)
    for d in fetch_server_deals():
        did = str(d.get("id") or "").strip()
        if not did:
            continue
        if did not in deals:
            deals[did] = {"customer": d.get("customer") or "", "owner": d.get("owner") or ""}
        else:
            if not deals[did].get("customer"):
                deals[did]["customer"] = d.get("customer") or ""
            if not deals[did].get("owner"):
                deals[did]["owner"] = d.get("owner") or ""
    dups = find_duplicate_customers(deals)
    if dups:
        print(f"ВНИМАНИЕ: найдено {len(dups)} дублей по имени клиента (строки по deal_id):")
        for c, items in list(dups.items())[:10]:
            print(f"  {items}")
    table = build_rows(deals, best)
    filled = sum(1 for r in table for lbl in PASSPORT_LABELS if r["cells"][lbl]["text"] != "—")
    print(f"Сделок: {len(table)}, заполненных ячеек (не «—»): {filled}")
    write_excel(table, dups, len(rows))
    print(f"Готово: {OUT}")


if __name__ == "__main__":
    main()
