import json, re, urllib.request
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
url = re.search(r'url:\s*"([^"]+)"', (ROOT/"js/gas-config.js").read_text(encoding="utf-8")).group(1)
current = json.loads(urllib.request.urlopen(url+"?action=get", timeout=120).read())["state"]

for name in ["truth_state_preview.json", "max_audit_state_preview.json"]:
    p = ROOT/"tools"/name
    if not p.exists():
        continue
    truth = json.loads(p.read_text(encoding="utf-8"))
    wb = openpyxl.load_workbook(Path(r"c:\Users\Данила\Downloads\инцидент 9-40-58.xlsx"), read_only=True, data_only=True)
    incident_ids = {str(r[2]).strip() for r in wb[wb.sheetnames[0]].iter_rows(min_row=2, values_only=True) if r and r[2]}
    import sys; sys.path.insert(0, str(ROOT/"tools"))
    from restore_max_audit_state import deal_field_count, score_sum
    cur = {d["id"]: d for d in current["deals"]}
    tru = {d["id"]: d for d in truth["deals"]}
    better = []
    for did in sorted(cur):
        if did in incident_ids:
            continue
        c, t = cur.get(did), tru.get(did)
        if not t:
            continue
        cf, tf = deal_field_count(c), deal_field_count(t)
        cs, ts = score_sum(c), score_sum(t)
        if tf > cf or ts > cs or json.dumps(t.get("techResearch"),sort_keys=True,default=str) != json.dumps(c.get("techResearch"),sort_keys=True,default=str):
            better.append((did, cf, tf, cs, ts))
    print(name, "non-incident deals differing:", len(better))
    for x in better[:15]:
        print(" ", x)
