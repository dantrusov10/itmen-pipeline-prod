import json
from pathlib import Path
import re, urllib.request
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
url = re.search(r'url:\s*"([^"]+)"', (ROOT/"js/gas-config.js").read_text(encoding="utf-8")).group(1)
current = json.loads(urllib.request.urlopen(url+"?action=get", timeout=120).read())["state"]
truth = json.loads((ROOT/"tools/truth_state_preview.json").read_text(encoding="utf-8"))

wb = openpyxl.load_workbook(Path(r"c:\Users\Данила\Downloads\инцидент 9-40-58.xlsx"), read_only=True, data_only=True)
incident_ids = {str(r[2]).strip() for r in wb[wb.sheetnames[0]].iter_rows(min_row=2, values_only=True) if r and r[2]}

sys_path = str(ROOT/"tools")
import sys; sys.path.insert(0, sys_path)
from restore_max_audit_state import deal_field_count, score_sum

cur = {d["id"]: d for d in current["deals"]}
tru = {d["id"]: d for d in truth["deals"]}

worse = []
for did in sorted(cur):
    if did in incident_ids:
        continue
    c, t = cur[did], tru.get(did)
    if not t:
        continue
    cf, tf = deal_field_count(c), deal_field_count(t)
    cs, ts = score_sum(c), score_sum(t)
    if tf > cf or ts > cs:
        worse.append((did, c.get("owner","")[:20], (c.get("customer") or "")[:30], cf, tf, cs, ts))

print(f"Non-incident deals where truth has MORE data than current: {len(worse)}")
for x in sorted(worse, key=lambda z: -(z[5]-z[4]))[:30]:
    print(f"  {x[0]} {x[2]} fields {x[3]}->{x[4]} score {x[5]}->{x[6]}")
