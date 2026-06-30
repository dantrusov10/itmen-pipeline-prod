import openpyxl
from collections import Counter
path = r"c:\Users\Данила\Downloads\инцидент 9-40-58.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
labels = Counter()
pct_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[2]:
        continue
    labels[str(row[6])] += 1
    if row[6] and "требован" in str(row[6]).lower():
        pct_rows.append((row[2], row[6], row[7], row[8], type(row[7]).__name__, type(row[8]).__name__))
print("field counts:", labels.most_common(10))
print("pct rows:", len(pct_rows))
for r in pct_rows[:15]:
    print(r)
