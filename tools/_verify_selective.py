import json, re, urllib.request, hashlib
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
url = re.search(r'url:\s*"([^"]+)"', (ROOT/"js/gas-config.js").read_text(encoding="utf-8")).group(1)

def deal_hash(d):
    return hashlib.md5(json.dumps(d, sort_keys=True, ensure_ascii=False, default=str).encode()).hexdigest()

# load excel deal ids
wb = openpyxl.load_workbook(Path(r"c:\Users\Данила\Downloads\инцидент 9-40-58.xlsx"), read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
incident_ids = {str(r[2]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r and r[2]}

# we need before state - use preview from recover script re-run
import subprocess
subprocess.run(["python", "recover_incident_0940_excel.py"], cwd=ROOT/"tools", check=True)

deals = json.loads(urllib.request.urlopen(url+"?action=get", timeout=120).read())["state"]["deals"]
not_in = [d for d in deals if d["id"] not in incident_ids]
in_inc = [d for d in deals if d["id"] in incident_ids]
print("incident deals", len(in_inc), "untouched deals", len(not_in))
# check pcts
pcts = [d.get("techResearch",{}).get("productRequirementsPct") for d in deals if d.get("techResearch",{}).get("productRequirementsPct") is not None]
print("product pct deals", len(pcts), "avg", round(sum(pcts)/len(pcts),1) if pcts else None)
bad = [v for v in pcts if v>100 or v<0]
print("bad pcts", bad)
